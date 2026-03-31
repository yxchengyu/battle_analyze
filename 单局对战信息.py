# 整体流程：
# 1. 查询单局对战明细。
# 2. 将结果中的 ID 翻译成中文名称。
# 3. 分别生成“基于武器类型拆分”“基于套装拆分”和“基于攻击部拆分”的透视结果。
# 4. 将三组透视连续导出到同一个 Excel 工作表，并同步到同一个飞书页签。
import json
import logging
from logging.handlers import RotatingFileHandler
from pathlib import Path
import re
import socket
import sys
import time
from typing import Any, Callable
from dataclasses import dataclass
from datetime import date, datetime, timedelta, timezone
from urllib.error import HTTPError, URLError
from urllib.parse import urlparse
from urllib.request import Request, urlopen


PROJECT_ROOT = Path(__file__).resolve().parents[1]

if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

try:
    import prestodb
except ImportError as exc:
    raise SystemExit(
        "当前环境未安装 prestodb，请先执行 `pip install prestodb` 后再运行脚本。"
    ) from exc

try:
    import openpyxl
    import pandas as pd
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError as exc:
    raise SystemExit(
        "当前环境未安装 pandas 或 openpyxl，请先执行 "
        "`pip install pandas openpyxl` 后再运行脚本。"
    ) from exc


LOGGER_NAME = "battle_analyze"
LOG_FILE_PATH = Path(__file__).with_name("battle_analyze.log")
CONFIG_FILE_PATH = Path(__file__).with_name("config.local.json")
CONFIG_EXAMPLE_FILE_PATH = Path(__file__).with_name("config.example.json")
logger = logging.getLogger(LOGGER_NAME)


@dataclass
class RuntimeConfig:
    presto_host: str
    presto_port: int
    presto_user: str
    presto_password: str
    presto_catalog: str
    presto_http_scheme: str
    feishu_app_id: str
    feishu_app_secret: str
    feishu_timeout_seconds: int
    feishu_report_wiki_url: str
    item_info_excel_path: Path


PRESTO_HOST = ""
PRESTO_PORT = 0
PRESTO_USER = ""
PRESTO_PASSWORD = ""
PRESTO_CATALOG = ""
PRESTO_HTTP_SCHEME = "https"
FEISHU_APP_ID = ""
FEISHU_APP_SECRET = ""
FEISHU_TIMEOUT_SECONDS = 30


# 报表基础配置。
ITEM_INFO_EXCEL_PATH = PROJECT_ROOT / "公共引用" / "英雄、武器、装备信息.xlsx"
FEISHU_REPORT_WIKI_URL = ""
HERO_USAGE_RATE_LOW_THRESHOLD = 0.05
HERO_USAGE_RATE_HIGH_THRESHOLD = 0.25
WEAPON_USAGE_RATE_LOW_THRESHOLD = 0.04
WEAPON_USAGE_RATE_HIGH_THRESHOLD = 0.20
ATTACK_PART_USAGE_RATE_ALERT_THRESHOLD = 0.02
SUMMARY_HIGH_FILL_COLOR = "#E06666"
SUMMARY_LOW_FILL_COLOR = "#93C47D"
REPORT_WORKSHEET_TITLE = "英雄武器攻击部奥义配件透视"
WEAPON_SECTION_TITLE = "基于武器类型拆分"
EQUIPMENT_SET_SECTION_TITLE = "基于套装拆分"
ATTACK_PART_SECTION_TITLE = "基于攻击部拆分"
WEAPON_ULT_SECTION_TITLE = "基于武器+奥义拆分"
WEAPON_BOOST_SECTION_TITLE = "基于配件拆分"
WEAPON_USAGE_DELTA_BLOCK_TITLE = "使用率变化"
WEAPON_EVACUATION_DELTA_BLOCK_TITLE = "撤离率变化"
SECTION_SIDE_BY_SIDE_GAP_COLUMNS = 2
WEEKLY_BASELINE_CACHE_PATH = Path(__file__).with_name("单局对战信息_weekly_cache.json")
VERSION_UPDATE_TIME_PATH = Path(__file__).with_name("version_update_times.json")
WEEKLY_BASELINE_REFRESH_WEEKDAY = 2
WEEKLY_BASELINE_START_WEEKDAY = 0
CHINA_TIMEZONE = timezone(timedelta(hours=8))
QUERY_DATE_VALUE = datetime.now(CHINA_TIMEZONE).date()
QUERY_DATE = QUERY_DATE_VALUE.strftime("%Y-%m-%d")
REPORT_DATE_LABEL = QUERY_DATE_VALUE.strftime("%m-%d")
CURRENT_PERIOD_LOWER_BOUND: str | None = None

TRANSLATE_COLUMNS = (
    "b_battle_begin_hero",
    "b_main_weapon_type",
    "b_main_weapon_variant",
    "b_main_weapon_ult",
    "b_main_weapon_boost",
    "b_sub_weapon_type",
    "b_sub_weapon_variant",
    "b_sub_weapon_ult",
    "b_sub_weapon_boost",
    "b_headset",
    "b_armor",
    "b_shoe",
)

MAIN_WEAPON_TYPE_ORDER = (
    "突击步枪",
    "脉冲双枪",
    "霰弹枪",
    "狙击枪",
    "能量弓",
    "手里剑",
    "飞轮枪",
    "雷切",
    "重锤",
    "镰刀",
    "长枪",
    "榴弹炮",
)
MAIN_WEAPON_TYPE_ORDER_MAP = {
    weapon_type: order_index
    for order_index, weapon_type in enumerate(MAIN_WEAPON_TYPE_ORDER)
}
EQUIPMENT_NAME_PREFIXES = ("粗糙", "优质", "顶级")
DEFAULT_BLOCK_DEFINITIONS = (
    {
        "title": "使用率",
        "data_key": "usage_rate_df",
        "percentage": True,
        "end_color": "F4B183",
        "highlight_low_usage": True,
        "color_mode": "global_single",
    },
    {
        "title": "撤离率",
        "data_key": "evacuation_rate_df",
        "percentage": True,
        "end_color": "FFF200",
        "highlight_low_usage": False,
        "color_mode": "global_single",
    },
    {
        "title": "平均带出价值增量",
        "data_key": "avg_inc_value_df",
        "percentage": False,
        "end_color": "9BC2E6",
        "highlight_low_usage": False,
        "color_mode": "global_single",
    },
    {
        "title": "使用次数",
        "data_key": "count_df",
        "percentage": False,
        "end_color": "E68A83",
        "highlight_low_usage": False,
        "color_mode": "global_single",
    },
)
WEAPON_EXTRA_BLOCK_DEFINITIONS = (
    {
        "title": "分均移动距离",
        "data_key": "avg_moved_per_minute_df",
        "percentage": False,
        "end_color": "C6E0B4",
        "highlight_low_usage": False,
        "color_mode": "global_single",
    },
    {
        "title": "击倒玩家数量",
        "data_key": "avg_player_down_df",
        "percentage": False,
        "end_color": "BDD7EE",
        "highlight_low_usage": False,
        "color_mode": "global_single",
    },
    {
        "title": "击杀怪物数",
        "data_key": "avg_monster_killed_df",
        "percentage": False,
        "end_color": "F8CBAD",
        "highlight_low_usage": False,
        "color_mode": "global_single",
    },
)
WEAPON_SECTION_BLOCK_DEFINITIONS = DEFAULT_BLOCK_DEFINITIONS[:-1] + WEAPON_EXTRA_BLOCK_DEFINITIONS + (
    DEFAULT_BLOCK_DEFINITIONS[-1],
)
SET_SECTION_BLOCK_DEFINITIONS = (
    DEFAULT_BLOCK_DEFINITIONS[0],
    DEFAULT_BLOCK_DEFINITIONS[1],
)
ALL_BLOCK_TITLES = {
    block_definition["title"]
    for block_definition in (
        DEFAULT_BLOCK_DEFINITIONS
        + WEAPON_EXTRA_BLOCK_DEFINITIONS
        + SET_SECTION_BLOCK_DEFINITIONS
    )
}
ALL_BLOCK_TITLES.update({WEAPON_USAGE_DELTA_BLOCK_TITLE, WEAPON_EVACUATION_DELTA_BLOCK_TITLE})
FEISHU_REQUEST_MAX_RETRIES = 3
FEISHU_VALUES_BATCH_SIZE = 4
WEEKLY_CACHE_VERSION = 1


def setup_logging() -> None:
    if logger.handlers:
        return

    formatter = logging.Formatter(
        "%(asctime)s [%(levelname)s] %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )
    file_handler = RotatingFileHandler(
        LOG_FILE_PATH,
        maxBytes=2 * 1024 * 1024,
        backupCount=3,
        encoding="utf-8",
    )
    console_handler = logging.StreamHandler(sys.stdout)

    file_handler.setFormatter(formatter)
    console_handler.setFormatter(formatter)

    logger.setLevel(logging.INFO)
    logger.addHandler(file_handler)
    logger.addHandler(console_handler)
    logger.propagate = False


def load_runtime_config(config_path: Path = CONFIG_FILE_PATH) -> RuntimeConfig:
    if not config_path.exists():
        raise SystemExit(
            f"未找到配置文件：{config_path}。"
            f"请先参考 {CONFIG_EXAMPLE_FILE_PATH.name} 创建 {config_path.name}。"
        )

    try:
        config_data = json.loads(config_path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise SystemExit(f"配置文件读取失败：{config_path}，{exc}") from exc

    try:
        presto_config = config_data["presto"]
        feishu_config = config_data["feishu"]
        report_config = config_data["report"]
    except KeyError as exc:
        raise SystemExit(f"配置文件缺少必要字段：{exc}") from exc

    item_info_excel_path = Path(str(report_config["item_info_excel_path"]).strip())
    if not item_info_excel_path.is_absolute():
        item_info_excel_path = (config_path.parent / item_info_excel_path).resolve()

    return RuntimeConfig(
        presto_host=str(presto_config["host"]).strip(),
        presto_port=int(presto_config["port"]),
        presto_user=str(presto_config["user"]).strip(),
        presto_password=str(presto_config["password"]).strip(),
        presto_catalog=str(presto_config["catalog"]).strip(),
        presto_http_scheme=str(presto_config.get("http_scheme", "https")).strip(),
        feishu_app_id=str(feishu_config["app_id"]).strip(),
        feishu_app_secret=str(feishu_config["app_secret"]).strip(),
        feishu_timeout_seconds=int(feishu_config.get("timeout_seconds", 30)),
        feishu_report_wiki_url=str(feishu_config["report_wiki_url"]).strip(),
        item_info_excel_path=item_info_excel_path,
    )


def apply_runtime_config(runtime_config: RuntimeConfig) -> None:
    global PRESTO_HOST
    global PRESTO_PORT
    global PRESTO_USER
    global PRESTO_PASSWORD
    global PRESTO_CATALOG
    global PRESTO_HTTP_SCHEME
    global FEISHU_APP_ID
    global FEISHU_APP_SECRET
    global FEISHU_TIMEOUT_SECONDS
    global FEISHU_REPORT_WIKI_URL
    global ITEM_INFO_EXCEL_PATH

    PRESTO_HOST = runtime_config.presto_host
    PRESTO_PORT = runtime_config.presto_port
    PRESTO_USER = runtime_config.presto_user
    PRESTO_PASSWORD = runtime_config.presto_password
    PRESTO_CATALOG = runtime_config.presto_catalog
    PRESTO_HTTP_SCHEME = runtime_config.presto_http_scheme
    FEISHU_APP_ID = runtime_config.feishu_app_id
    FEISHU_APP_SECRET = runtime_config.feishu_app_secret
    FEISHU_TIMEOUT_SECONDS = runtime_config.feishu_timeout_seconds
    FEISHU_REPORT_WIKI_URL = runtime_config.feishu_report_wiki_url
    ITEM_INFO_EXCEL_PATH = runtime_config.item_info_excel_path


def set_runtime_query_context(query_date_value: date) -> None:
    global QUERY_DATE_VALUE
    global QUERY_DATE
    global REPORT_DATE_LABEL
    global CURRENT_PERIOD_LOWER_BOUND

    QUERY_DATE_VALUE = query_date_value
    QUERY_DATE = query_date_value.strftime("%Y-%m-%d")
    REPORT_DATE_LABEL = query_date_value.strftime("%m-%d")
    CURRENT_PERIOD_LOWER_BOUND = build_current_period_lower_bound(query_date_value)


def resolve_query_date() -> date:
    raw_value = sys.argv[1] if len(sys.argv) > 1 else ""
    if not raw_value:
        return datetime.now().date()

    normalized_value = raw_value.strip().replace(".", "-").replace("/", "-")
    try:
        return datetime.strptime(normalized_value, "%Y-%m-%d").date()
    except ValueError as exc:
        raise SystemExit("查询日期格式错误，请使用 YYYY-MM-DD，例如 2026-03-27。") from exc


def load_version_update_times(config_path: Path) -> dict[str, str]:
    if not config_path.exists():
        return {}

    try:
        config_data = json.loads(config_path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise SystemExit(f"版本更新时间配置文件读取失败：{config_path}，{exc}") from exc

    if not isinstance(config_data, dict):
        raise SystemExit(f"版本更新时间配置文件格式错误：{config_path}，顶层必须是对象。")

    normalized_config: dict[str, str] = {}
    for date_text, time_text in config_data.items():
        normalized_date = str(date_text).strip()
        normalized_time = str(time_text).strip()
        try:
            datetime.strptime(normalized_date, "%Y-%m-%d")
            datetime.strptime(normalized_time, "%H:%M:%S")
        except ValueError as exc:
            raise SystemExit(
                f"版本更新时间配置格式错误：{normalized_date} -> {normalized_time}，"
                "请使用 YYYY-MM-DD: HH:MM:SS"
            ) from exc
        normalized_config[normalized_date] = normalized_time

    return normalized_config


def build_current_period_lower_bound(query_date_value: date) -> str | None:
    query_date_text = query_date_value.strftime("%Y-%m-%d")
    version_update_times = load_version_update_times(VERSION_UPDATE_TIME_PATH)
    update_time_text = version_update_times.get(query_date_text)

    if not update_time_text:
        return None

    update_local_datetime = datetime.strptime(
        f"{query_date_text} {update_time_text}",
        "%Y-%m-%d %H:%M:%S",
    ).replace(tzinfo=CHINA_TIMEZONE)
    update_utc_datetime = update_local_datetime.astimezone(timezone.utc)
    return update_utc_datetime.strftime("%Y-%m-%d %H:%M:%S")


# 主查询：抽取单局对战明细，后续透视统计都基于这份结果做。
def build_sql_query(
    query_dates: list[str],
    time_lower_bound: str | None = None,
) -> str:
    date_text = ", ".join(f"'{query_date}'" for query_date in query_dates)
    time_filter_sql = ""
    if time_lower_bound:
        time_filter_sql = f'\n    AND a."time" > TIMESTAMP \'{time_lower_bound}\''

    return f"""
SELECT
    a.roleid AS a_roleid,
    a.zoneid AS a_zoneid,
    a.pvpmode AS a_pvpmode,
    b.battle_begin_hero AS b_battle_begin_hero,
    -- 分号分隔拆分，Presto 数组下标从 1 开始
    element_at(split(element_at(split(b.battle_begin_item, chr(59)), 1), ','), 1) AS b_main_weapon_type,
    element_at(split(element_at(split(b.battle_begin_item, chr(59)), 1), ','), 3) AS b_main_weapon_variant,
    element_at(split(element_at(split(b.battle_begin_item, chr(59)), 1), ','), 4) AS b_main_weapon_ult,
    element_at(split(element_at(split(b.battle_begin_item, chr(59)), 1), ','), 5) AS b_main_weapon_boost,

    element_at(split(element_at(split(b.battle_begin_item, chr(59)), 2), ','), 1) AS b_sub_weapon_type,
    element_at(split(element_at(split(b.battle_begin_item, chr(59)), 2), ','), 3) AS b_sub_weapon_variant,
    element_at(split(element_at(split(b.battle_begin_item, chr(59)), 2), ','), 4) AS b_sub_weapon_ult,
    element_at(split(element_at(split(b.battle_begin_item, chr(59)), 2), ','), 5) AS b_sub_weapon_boost,

    element_at(split(b.battle_begin_item, chr(59)), 3) AS b_headset,
    element_at(split(b.battle_begin_item, chr(59)), 4) AS b_armor,
    element_at(split(b.battle_begin_item, chr(59)), 5) AS b_shoe,
    b.equip_value AS b_equip_value,
    a.battleid AS a_battleid,
    a.battletime AS a_battletime,
    a.total_value_nokey AS a_total_value_nokey,
    a.evacuatedsucc AS a_evacuatedsucc,
    a.inc_value AS a_inc_value,
    a.failed_mod AS a_failed_mod,
    a.moved AS a_moved,
    a.player_down AS a_player_down,
    a.monster_killed AS a_monster_killed,
    a.carry_out_value AS a_carry_out_value

FROM sgame_ods.battleserver_battle_end a
JOIN sgame_ods.battleserver_battle_begin b
    ON a.battleid = b.battleid
    AND a.roleid = b.roleid
WHERE a.logymd IN ({date_text})
    AND b.logymd IN ({date_text}){time_filter_sql}
    AND a.real_player_count >= 6
    AND a.level > 15
    AND a.pvpmode != 109
ORDER BY a."time" DESC
"""


class FeishuApiError(Exception):
    """飞书接口调用异常。"""


def get_connection() -> prestodb.dbapi.Connection:
    # 创建 Presto 连接，所有查询都通过这里统一接入。
    return prestodb.dbapi.connect(
        host=PRESTO_HOST,
        port=PRESTO_PORT,
        user=PRESTO_USER,
        catalog=PRESTO_CATALOG,
        http_scheme=PRESTO_HTTP_SCHEME,
        auth=prestodb.auth.BasicAuthentication(PRESTO_USER, PRESTO_PASSWORD),
    )


def build_available_output_path(output_file_path: Path) -> Path:
    if not output_file_path.exists():
        return output_file_path

    try:
        with output_file_path.open("ab"):
            return output_file_path
    except PermissionError:
        timestamp_text = datetime.now().strftime("%Y%m%d_%H%M%S")
        return output_file_path.with_name(
            f"{output_file_path.stem}_{timestamp_text}{output_file_path.suffix}"
        )


def execute_query(sql: str) -> tuple[list[str], list[tuple[Any, ...]]]:
    connection = get_connection()
    cursor = connection.cursor()

    try:
        cursor.execute(sql)
        column_names = [column[0] for column in cursor.description]
        rows = cursor.fetchall()
        return column_names, rows
    finally:
        cursor.close()
        connection.close()


def build_weekly_baseline_window(anchor_date: date) -> dict[str, Any]:
    days_since_wednesday = (anchor_date.weekday() - WEEKLY_BASELINE_REFRESH_WEEKDAY) % 7
    baseline_end = anchor_date - timedelta(days=days_since_wednesday)
    baseline_start = baseline_end - timedelta(
        days=WEEKLY_BASELINE_REFRESH_WEEKDAY - WEEKLY_BASELINE_START_WEEKDAY
    )
    query_dates = [
        (baseline_start + timedelta(days=day_offset)).strftime("%Y-%m-%d")
        for day_offset in range((baseline_end - baseline_start).days + 1)
    ]
    return {
        "week_key": baseline_start.strftime("%Y-%m-%d"),
        "start_date": baseline_start,
        "end_date": baseline_end,
        "query_dates": query_dates,
    }


def load_weekly_baseline_cache(cache_path: Path) -> dict[str, Any]:
    if not cache_path.exists():
        return {}

    try:
        cache_data = json.loads(cache_path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}

    if not isinstance(cache_data, dict):
        return {}
    if cache_data.get("version") != WEEKLY_CACHE_VERSION:
        return {}

    return cache_data


def save_weekly_baseline_cache(cache_path: Path, cache_data: dict[str, Any]) -> None:
    cache_data = cache_data.copy()
    cache_data["version"] = WEEKLY_CACHE_VERSION
    cache_data["updated_at"] = datetime.now().replace(microsecond=0).isoformat()
    cache_path.write_text(
        json.dumps(cache_data, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def serialize_dataframe(data_frame: pd.DataFrame) -> dict[str, Any]:
    serialized_df = data_frame.copy().astype(object)
    serialized_df = serialized_df.where(pd.notna(serialized_df), None)
    return {
        "index": [str(index_value) for index_value in serialized_df.index.tolist()],
        "columns": [str(column_name) for column_name in serialized_df.columns.tolist()],
        "data": serialized_df.values.tolist(),
    }


def deserialize_dataframe(payload: dict[str, Any]) -> pd.DataFrame:
    if not payload:
        return pd.DataFrame()

    return pd.DataFrame(
        payload.get("data", []),
        index=payload.get("index", []),
        columns=payload.get("columns", []),
    )


def align_delta_frames(
    current_df: pd.DataFrame,
    baseline_df: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    aligned_index = list(dict.fromkeys(current_df.index.tolist() + baseline_df.index.tolist()))
    aligned_columns = list(dict.fromkeys(current_df.columns.tolist() + baseline_df.columns.tolist()))
    if "*" in aligned_index:
        aligned_index = [label for label in aligned_index if label != "*"] + ["*"]
    if "*" in aligned_columns:
        aligned_columns = [label for label in aligned_columns if label != "*"] + ["*"]
    current_aligned_df = current_df.reindex(index=aligned_index, columns=aligned_columns)
    baseline_aligned_df = baseline_df.reindex(index=aligned_index, columns=aligned_columns)
    return current_aligned_df, baseline_aligned_df


def build_delta_dataframe(
    current_df: pd.DataFrame,
    baseline_df: pd.DataFrame,
) -> pd.DataFrame:
    current_aligned_df, baseline_aligned_df = align_delta_frames(current_df, baseline_df)
    delta_df = current_aligned_df.astype(float) - baseline_aligned_df.astype(float)
    return delta_df


def build_weapon_baseline_delta_tables(
    pivot_source_df: pd.DataFrame,
    equipment_name_to_set_mapping: dict[str, str],
    set_name_candidates: list[str],
    anchor_date: date,
) -> tuple[pd.DataFrame, pd.DataFrame, dict[str, Any]]:
    baseline_window = build_weekly_baseline_window(anchor_date)
    cache_data = load_weekly_baseline_cache(WEEKLY_BASELINE_CACHE_PATH)
    cache_week_key = str(cache_data.get("week_key", "")).strip()

    if cache_week_key == baseline_window["week_key"]:
        baseline_usage_rate_df = deserialize_dataframe(cache_data.get("weapon_usage_rate_df", {}))
        baseline_evacuation_rate_df = deserialize_dataframe(cache_data.get("weapon_evacuation_rate_df", {}))
        if not baseline_usage_rate_df.empty and not baseline_evacuation_rate_df.empty:
            logger.info(
                "已命中周基准缓存："
                f"{baseline_window['start_date']:%m-%d}~{baseline_window['end_date']:%m-%d}"
            )
        else:
            cache_week_key = ""

    if cache_week_key != baseline_window["week_key"]:
        logger.info(
            "开始拉取周基准数据："
            f"{baseline_window['start_date']:%Y-%m-%d} ~ {baseline_window['end_date']:%Y-%m-%d}"
        )
        baseline_sql = build_sql_query(baseline_window["query_dates"])
        baseline_column_names, baseline_rows = execute_query(baseline_sql)
        baseline_result_df = pd.DataFrame(baseline_rows, columns=baseline_column_names)
        baseline_result_df = normalize_weapon_source_columns(baseline_result_df)
        id_name_mapping = load_item_name_mapping(ITEM_INFO_EXCEL_PATH)
        baseline_result_df = translate_result_dataframe(baseline_result_df, id_name_mapping)
        baseline_pivot_source_df = prepare_pivot_source_dataframe(
            baseline_result_df,
            equipment_name_to_set_mapping,
            set_name_candidates,
        )
        baseline_tables = build_hero_weapon_pivot_tables(
            baseline_pivot_source_df,
            "b_main_weapon_type",
            column_sort_key=build_weapon_sort_key,
        )
        baseline_usage_rate_df = baseline_tables["usage_rate_df"]
        baseline_evacuation_rate_df = baseline_tables["evacuation_rate_df"]
        save_weekly_baseline_cache(
            WEEKLY_BASELINE_CACHE_PATH,
            {
                "week_key": baseline_window["week_key"],
                "start_date": baseline_window["start_date"].isoformat(),
                "end_date": baseline_window["end_date"].isoformat(),
                "weapon_usage_rate_df": serialize_dataframe(baseline_usage_rate_df),
                "weapon_evacuation_rate_df": serialize_dataframe(baseline_evacuation_rate_df),
            },
        )

    current_tables = build_hero_weapon_pivot_tables(
        pivot_source_df,
        "b_main_weapon_type",
        column_sort_key=build_weapon_sort_key,
    )
    usage_delta_df = build_delta_dataframe(
        current_tables["usage_rate_df"],
        baseline_usage_rate_df,
    )
    evacuation_delta_df = build_delta_dataframe(
        current_tables["evacuation_rate_df"],
        baseline_evacuation_rate_df,
    )
    return usage_delta_df, evacuation_delta_df, baseline_window


def _request_json(
    method: str,
    url: str,
    headers: dict[str, str] | None = None,
    body: dict[str, Any] | None = None,
) -> dict[str, Any]:
    request_headers = headers.copy() if headers else {}
    request_body = None

    if body is not None:
        request_body = json.dumps(body, ensure_ascii=False).encode("utf-8")
        request_headers["Content-Type"] = "application/json; charset=utf-8"
    last_error: Exception | None = None

    for attempt in range(1, FEISHU_REQUEST_MAX_RETRIES + 1):
        request = Request(url=url, data=request_body, headers=request_headers, method=method)

        try:
            with urlopen(request, timeout=FEISHU_TIMEOUT_SECONDS) as response:
                response_text = response.read().decode("utf-8")
            break
        except HTTPError as exc:
            error_text = exc.read().decode("utf-8", errors="ignore")
            raise FeishuApiError(f"HTTP 错误：{exc.code}，响应内容：{error_text}") from exc
        except (URLError, TimeoutError, socket.timeout) as exc:
            last_error = exc
            if attempt >= FEISHU_REQUEST_MAX_RETRIES:
                raise FeishuApiError(
                    f"网络请求失败，已重试 {FEISHU_REQUEST_MAX_RETRIES} 次：{exc}"
                ) from exc
            logger.warning(
                f"飞书请求失败，第 {attempt} 次重试中：{exc} "
                f"(URL: {url})"
            )
            time.sleep(attempt)
    else:
        raise FeishuApiError(f"网络请求失败：{last_error}")

    try:
        return json.loads(response_text)
    except json.JSONDecodeError as exc:
        raise FeishuApiError(f"响应不是合法的 JSON：{response_text}") from exc


def upload_value_ranges_in_batches(
    access_token: str,
    spreadsheet_token: str,
    value_ranges: list[dict[str, Any]],
    progress_label: str,
) -> None:
    if not value_ranges:
        return

    url = (
        f"https://open.feishu.cn/open-apis/sheets/v2/spreadsheets/"
        f"{spreadsheet_token}/values_batch_update"
    )
    headers = {"Authorization": f"Bearer {access_token}"}
    total_batches = (len(value_ranges) + FEISHU_VALUES_BATCH_SIZE - 1) // FEISHU_VALUES_BATCH_SIZE

    for batch_index, start_index in enumerate(range(0, len(value_ranges), FEISHU_VALUES_BATCH_SIZE), start=1):
        batch_ranges = value_ranges[start_index:start_index + FEISHU_VALUES_BATCH_SIZE]
        logger.info(
            f"飞书写入中：{progress_label}，第 {batch_index}/{total_batches} 批，"
            f"本批 {len(batch_ranges)} 个区域"
        )
        result = _request_json(
            "POST",
            url,
            headers=headers,
            body={"valueRanges": batch_ranges},
        )

        if result.get("code") != 0:
            raise FeishuApiError(f"写入飞书页签失败：{result.get('msg', '未知错误')}")


def update_feishu_dimension_range(
    access_token: str,
    spreadsheet_token: str,
    dimension: dict[str, Any],
    dimension_properties: dict[str, Any],
) -> None:
    url = (
        f"https://open.feishu.cn/open-apis/sheets/v2/spreadsheets/"
        f"{spreadsheet_token}/dimension_range"
    )
    result = _request_json(
        "PUT",
        url,
        headers={"Authorization": f"Bearer {access_token}"},
        body={
            "dimension": dimension,
            "dimensionProperties": dimension_properties,
        },
    )

    if result.get("code") != 0:
        raise FeishuApiError(f"更新飞书行列属性失败：{result.get('msg', '未知错误')}")


def get_feishu_access_token() -> str:
    url = "https://open.feishu.cn/open-apis/auth/v3/tenant_access_token/internal"
    body = {
        "app_id": FEISHU_APP_ID,
        "app_secret": FEISHU_APP_SECRET,
    }
    result = _request_json("POST", url, body=body)

    if result.get("code") != 0:
        raise FeishuApiError(f"获取飞书访问令牌失败：{result.get('msg', '未知错误')}")

    access_token = result.get("tenant_access_token")
    if not access_token:
        raise FeishuApiError("获取飞书访问令牌失败：响应中未返回 tenant_access_token。")

    return access_token


def extract_wiki_token(wiki_url: str) -> str:
    parsed_url = urlparse(wiki_url)
    wiki_token = parsed_url.path.rstrip("/").split("/")[-1]

    if not wiki_token:
        raise FeishuApiError("飞书 Wiki 链接中缺少 wiki_token。")

    return wiki_token


def get_feishu_sheet_token(access_token: str, wiki_url: str) -> str:
    wiki_token = extract_wiki_token(wiki_url)
    url = f"https://open.feishu.cn/open-apis/wiki/v2/spaces/get_node?token={wiki_token}"
    headers = {"Authorization": f"Bearer {access_token}"}
    result = _request_json("GET", url, headers=headers)

    if result.get("code") != 0:
        raise FeishuApiError(f"获取 Wiki 节点信息失败：{result.get('msg', '未知错误')}")

    node = result.get("data", {}).get("node", {})
    obj_type = node.get("obj_type")
    obj_token = node.get("obj_token")

    if obj_type != "sheet" or not obj_token:
        raise FeishuApiError(f"当前链接不是电子表格节点，obj_type={obj_type}")

    return obj_token


def create_feishu_sheet(
    access_token: str,
    spreadsheet_token: str,
    sheet_title: str,
) -> str:
    url = (
        f"https://open.feishu.cn/open-apis/sheets/v2/spreadsheets/"
        f"{spreadsheet_token}/sheets_batch_update"
    )
    headers = {"Authorization": f"Bearer {access_token}"}
    body = {
        "requests": [
            {
                "addSheet": {
                    "properties": {
                        "title": sheet_title,
                    }
                }
            }
        ]
    }
    result = _request_json("POST", url, headers=headers, body=body)

    if result.get("code") != 0:
        raise FeishuApiError(f"创建飞书页签失败：{result.get('msg', '未知错误')}")

    replies = result.get("data", {}).get("replies", [])
    if not replies:
        raise FeishuApiError("创建飞书页签失败：响应中未返回 replies。")

    properties = replies[0].get("addSheet", {}).get("properties", {})
    sheet_id = properties.get("sheetId")
    if not sheet_id:
        raise FeishuApiError("创建飞书页签失败：响应中未返回 sheetId。")

    return sheet_id


def build_excel_column_name(column_index: int) -> str:
    letters: list[str] = []
    current_index = column_index

    while current_index > 0:
        current_index, remainder = divmod(current_index - 1, 26)
        letters.append(chr(ord("A") + remainder))

    return "".join(reversed(letters))


def batch_update_feishu_styles(
    access_token: str,
    spreadsheet_token: str,
    style_requests: list[dict[str, Any]],
) -> None:
    if not style_requests:
        return

    url = (
        f"https://open.feishu.cn/open-apis/sheets/v2/spreadsheets/"
        f"{spreadsheet_token}/styles_batch_update"
    )
    headers = {"Authorization": f"Bearer {access_token}"}

    chunk_size = 200
    for start_index in range(0, len(style_requests), chunk_size):
        chunk_requests = style_requests[start_index:start_index + chunk_size]
        body = {"data": chunk_requests}
        result = _request_json("PUT", url, headers=headers, body=body)

        if result.get("code") != 0:
            raise FeishuApiError(f"批量设置飞书样式失败：{result.get('msg', '未知错误')}")


def build_feishu_style_request(
    ranges: list[str],
    style: dict[str, Any],
) -> dict[str, Any]:
    return {
        "ranges": ranges,
        "style": style,
    }


def build_alignment_style_request(cell_ranges: list[str]) -> dict[str, Any]:
    return build_feishu_style_request(
        cell_ranges,
        {
            "hAlign": 1,
            "vAlign": 1,
            "clean": False,
        },
    )


def build_left_alignment_style_request(cell_ranges: list[str]) -> dict[str, Any]:
    return build_feishu_style_request(
        cell_ranges,
        {
            "hAlign": 0,
            "vAlign": 1,
            "clean": False,
        },
    )


def build_percentage_style_request(cell_ranges: list[str]) -> dict[str, Any]:
    return build_feishu_style_request(
        cell_ranges,
        {
            "formatter": "0.00%",
            "clean": False,
        },
    )


def build_background_style_request(
    cell_ranges: list[str],
    back_color: str,
) -> dict[str, Any]:
    return build_feishu_style_request(
        cell_ranges,
        {
            "backColor": back_color,
            "clean": False,
        },
    )


def build_bold_style_request(cell_ranges: list[str]) -> dict[str, Any]:
    return build_feishu_style_request(
        cell_ranges,
        {
            "font": {
                "bold": True,
            },
            "clean": False,
        },
    )


def build_output_excel_path(report_name: str) -> Path:
    suffix = report_name.strip()
    return Path(__file__).with_name(f"{REPORT_DATE_LABEL}{suffix}.xlsx")


def build_feishu_sheet_title(report_name: str) -> str:
    return f"{REPORT_DATE_LABEL}{report_name.strip()}"


def hex_to_rgb(color: str) -> tuple[int, int, int]:
    normalized_color = color.lstrip("#")
    return (
        int(normalized_color[0:2], 16),
        int(normalized_color[2:4], 16),
        int(normalized_color[4:6], 16),
    )


def rgb_to_hex(rgb_color: tuple[int, int, int]) -> str:
    return "#{:02X}{:02X}{:02X}".format(*rgb_color)


def should_bold_label(value: Any) -> bool:
    if value is None:
        return False

    text = str(value).strip()
    if not text:
        return False
    if text == "*":
        return True

    return any("\u4e00" <= char <= "\u9fff" for char in text)


def interpolate_hex_color(start_color: str, end_color: str, ratio: float) -> str:
    clamped_ratio = min(max(ratio, 0.0), 1.0)
    start_rgb = hex_to_rgb(start_color)
    end_rgb = hex_to_rgb(end_color)
    interpolated_rgb = tuple(
        round(start_channel + (end_channel - start_channel) * clamped_ratio)
        for start_channel, end_channel in zip(start_rgb, end_rgb)
    )
    return rgb_to_hex(interpolated_rgb)


def build_pivot_block_values(
    title: str,
    pivot_df: pd.DataFrame,
    percentage: bool = False,
) -> list[list[Any]]:
    block_values: list[list[Any]] = [[title]]
    header_row = [""] + pivot_df.columns.tolist()
    block_values.append(header_row)

    for row_name in pivot_df.index:
        row_values: list[Any] = [row_name]
        for column_name in pivot_df.columns:
            value = pivot_df.loc[row_name, column_name]
            if pd.isna(value):
                row_values.append("")
            elif percentage:
                row_values.append(round(float(value), 4))
            else:
                try:
                    normalized_value = value.item()
                except AttributeError:
                    normalized_value = value
                if isinstance(normalized_value, int):
                    row_values.append(normalized_value)
                elif isinstance(normalized_value, float):
                    row_values.append(round(normalized_value, 2))
                else:
                    row_values.append(normalized_value)
        block_values.append(row_values)

    return block_values


def build_value_range(
    sheet_id: str,
    start_row: int,
    values: list[list[Any]],
    start_column: int = 1,
) -> dict[str, Any]:
    max_columns = max(len(row) for row in values)
    normalized_values = [row + [""] * (max_columns - len(row)) for row in values]
    start_column_name = build_excel_column_name(start_column)
    last_column_name = build_excel_column_name(start_column + max_columns - 1)
    last_row = start_row + len(normalized_values) - 1
    return {
        "range": f"{sheet_id}!{start_column_name}{start_row}:{last_column_name}{last_row}",
        "values": normalized_values,
    }


def iter_matrix_labels(
    pivot_df: pd.DataFrame,
    include_totals: bool,
) -> tuple[list[str], list[str]]:
    row_labels = [row_name for row_name in pivot_df.index.tolist() if include_totals or row_name != "*"]
    column_labels = [
        column_name for column_name in pivot_df.columns.tolist() if include_totals or column_name != "*"
    ]
    return row_labels, column_labels


def apply_single_hue_fill_to_excel(
    worksheet: openpyxl.worksheet.worksheet.Worksheet,
    pivot_df: pd.DataFrame,
    start_row: int,
    start_column: int,
    end_color: str,
    include_totals: bool = False,
) -> None:
    row_labels, column_labels = iter_matrix_labels(pivot_df, include_totals=include_totals)
    if not row_labels or not column_labels:
        return

    numeric_values = [
        float(pivot_df.loc[row_name, column_name])
        for row_name in row_labels
        for column_name in column_labels
        if not pd.isna(pivot_df.loc[row_name, column_name])
    ]
    if not numeric_values:
        return

    min_value = min(numeric_values)
    max_value = max(numeric_values)

    for row_offset, row_name in enumerate(row_labels):
        excel_row = start_row + 2 + row_offset
        for column_offset, column_name in enumerate(column_labels, start=start_column + 1):
            value = pivot_df.loc[row_name, column_name]
            if pd.isna(value):
                continue

            numeric_value = float(value)
            if max_value == min_value:
                ratio = 1.0 if numeric_value > 0 else 0.0
            else:
                ratio = (numeric_value - min_value) / (max_value - min_value)

            if ratio <= 0:
                continue

            worksheet.cell(row=excel_row, column=column_offset).fill = PatternFill(
                fill_type="solid",
                fgColor=interpolate_hex_color("#FFFFFF", end_color, ratio).lstrip("#"),
            )


def apply_row_based_single_hue_fill_to_excel(
    worksheet: openpyxl.worksheet.worksheet.Worksheet,
    pivot_df: pd.DataFrame,
    start_row: int,
    start_column: int,
    end_color: str,
    include_totals: bool = False,
) -> None:
    row_labels, column_labels = iter_matrix_labels(pivot_df, include_totals=include_totals)
    if not row_labels or not column_labels:
        return

    for row_offset, row_name in enumerate(row_labels):
        numeric_values = [
            float(pivot_df.loc[row_name, column_name])
            for column_name in column_labels
            if not pd.isna(pivot_df.loc[row_name, column_name])
        ]
        if not numeric_values:
            continue

        row_min = min(numeric_values)
        row_max = max(numeric_values)
        excel_row = start_row + 2 + row_offset

        for column_offset, column_name in enumerate(column_labels, start=start_column + 1):
            value = pivot_df.loc[row_name, column_name]
            if pd.isna(value):
                continue

            numeric_value = float(value)
            if row_max == row_min:
                ratio = 1.0 if numeric_value > 0 else 0.0
            else:
                ratio = (numeric_value - row_min) / (row_max - row_min)

            if ratio <= 0:
                continue

            worksheet.cell(row=excel_row, column=column_offset).fill = PatternFill(
                fill_type="solid",
                fgColor=interpolate_hex_color("#FFFFFF", end_color, ratio).lstrip("#"),
            )


def apply_row_based_centered_fill_to_excel(
    worksheet: openpyxl.worksheet.worksheet.Worksheet,
    pivot_df: pd.DataFrame,
    start_row: int,
    start_column: int,
    positive_end_color: str,
    negative_end_color: str,
    include_totals: bool = False,
) -> None:
    row_labels, column_labels = iter_matrix_labels(pivot_df, include_totals=include_totals)
    if not row_labels or not column_labels:
        return

    for row_offset, row_name in enumerate(row_labels):
        if "*" in pivot_df.columns and not pd.isna(pivot_df.loc[row_name, "*"]):
            center_value = float(pivot_df.loc[row_name, "*"])
        else:
            center_value = None

        numeric_values = [
            float(pivot_df.loc[row_name, column_name])
            for column_name in column_labels
            if not pd.isna(pivot_df.loc[row_name, column_name])
        ]
        if not numeric_values:
            continue
        if center_value is None:
            center_value = sum(numeric_values) / len(numeric_values)

        row_min = min(numeric_values)
        row_max = max(numeric_values)
        excel_row = start_row + 2 + row_offset

        for column_offset, column_name in enumerate(column_labels, start=start_column + 1):
            value = pivot_df.loc[row_name, column_name]
            if pd.isna(value):
                continue

            numeric_value = float(value)
            if numeric_value >= center_value:
                denominator = max(row_max - center_value, 0.0)
                ratio = 0.0 if denominator == 0 else (numeric_value - center_value) / denominator
                target_color = positive_end_color
            else:
                denominator = max(center_value - row_min, 0.0)
                ratio = 0.0 if denominator == 0 else (center_value - numeric_value) / denominator
                target_color = negative_end_color

            if ratio <= 0:
                continue

            worksheet.cell(row=excel_row, column=column_offset).fill = PatternFill(
                fill_type="solid",
                fgColor=interpolate_hex_color("#FFFFFF", target_color, ratio).lstrip("#"),
            )


def apply_diverging_fill_to_excel(
    worksheet: openpyxl.worksheet.worksheet.Worksheet,
    pivot_df: pd.DataFrame,
    start_row: int,
    start_column: int,
    positive_end_color: str,
    negative_end_color: str,
    include_totals: bool = True,
) -> None:
    row_labels, column_labels = iter_matrix_labels(pivot_df, include_totals=include_totals)
    if not row_labels or not column_labels:
        return

    positive_max = 0.0
    negative_max = 0.0
    for row_name in row_labels:
        for column_name in column_labels:
            value = pivot_df.loc[row_name, column_name]
            if pd.isna(value):
                continue
            numeric_value = float(value)
            if numeric_value > 0:
                positive_max = max(positive_max, numeric_value)
            elif numeric_value < 0:
                negative_max = max(negative_max, abs(numeric_value))

    if positive_max == 0 and negative_max == 0:
        return

    for row_offset, row_name in enumerate(row_labels):
        excel_row = start_row + 2 + row_offset
        for column_offset, column_name in enumerate(column_labels, start=start_column + 1):
            value = pivot_df.loc[row_name, column_name]
            if pd.isna(value):
                continue

            numeric_value = float(value)
            if numeric_value > 0 and positive_max > 0:
                ratio = numeric_value / positive_max
                target_color = positive_end_color
            elif numeric_value < 0 and negative_max > 0:
                ratio = abs(numeric_value) / negative_max
                target_color = negative_end_color
            else:
                continue

            worksheet.cell(row=excel_row, column=column_offset).fill = PatternFill(
                fill_type="solid",
                fgColor=interpolate_hex_color("#FFFFFF", target_color, ratio).lstrip("#"),
            )


def apply_centered_summary_fill_to_excel(
    worksheet: openpyxl.worksheet.worksheet.Worksheet,
    pivot_df: pd.DataFrame,
    start_row: int,
    start_column: int,
    positive_end_color: str,
    negative_end_color: str,
) -> None:
    if "*" not in pivot_df.index or "*" not in pivot_df.columns:
        return

    grand_total = pivot_df.loc["*", "*"]
    if pd.isna(grand_total):
        return
    center_value = float(grand_total)

    total_column_index = pivot_df.columns.get_loc("*") + start_column + 1
    total_row_index = start_row + 2 + pivot_df.index.get_loc("*")

    hero_summary_values = [
        float(pivot_df.loc[row_name, "*"])
        for row_name in pivot_df.index
        if row_name != "*" and not pd.isna(pivot_df.loc[row_name, "*"])
    ]
    weapon_summary_values = [
        float(pivot_df.loc["*", column_name])
        for column_name in pivot_df.columns
        if column_name != "*" and not pd.isna(pivot_df.loc["*", column_name])
    ]

    hero_min = min(hero_summary_values) if hero_summary_values else center_value
    hero_max = max(hero_summary_values) if hero_summary_values else center_value
    weapon_min = min(weapon_summary_values) if weapon_summary_values else center_value
    weapon_max = max(weapon_summary_values) if weapon_summary_values else center_value

    for row_offset, row_name in enumerate(pivot_df.index):
        if row_name == "*":
            continue
        value = pivot_df.loc[row_name, "*"]
        if pd.isna(value):
            continue
        numeric_value = float(value)
        if numeric_value >= center_value:
            denominator = max(hero_max - center_value, 0.0)
            ratio = 0.0 if denominator == 0 else (numeric_value - center_value) / denominator
            target_color = positive_end_color
        else:
            denominator = max(center_value - hero_min, 0.0)
            ratio = 0.0 if denominator == 0 else (center_value - numeric_value) / denominator
            target_color = negative_end_color
        if ratio <= 0:
            continue
        excel_row = start_row + 2 + row_offset
        worksheet.cell(row=excel_row, column=total_column_index).fill = PatternFill(
            fill_type="solid",
            fgColor=interpolate_hex_color("#FFFFFF", target_color, ratio).lstrip("#"),
        )

    for column_offset, column_name in enumerate(pivot_df.columns, start=start_column + 1):
        if column_name == "*":
            continue
        value = pivot_df.loc["*", column_name]
        if pd.isna(value):
            continue
        numeric_value = float(value)
        if numeric_value >= center_value:
            denominator = max(weapon_max - center_value, 0.0)
            ratio = 0.0 if denominator == 0 else (numeric_value - center_value) / denominator
            target_color = positive_end_color
        else:
            denominator = max(center_value - weapon_min, 0.0)
            ratio = 0.0 if denominator == 0 else (center_value - numeric_value) / denominator
            target_color = negative_end_color
        if ratio <= 0:
            continue
        worksheet.cell(row=total_row_index, column=column_offset).fill = PatternFill(
            fill_type="solid",
            fgColor=interpolate_hex_color("#FFFFFF", target_color, ratio).lstrip("#"),
        )


def apply_excel_fill_to_block(
    worksheet: openpyxl.worksheet.worksheet.Worksheet,
    pivot_df: pd.DataFrame,
    start_row: int,
    start_column: int,
    block_definition: dict[str, Any],
) -> None:
    color_mode = block_definition.get("color_mode", "global_single")
    if color_mode == "weapon_usage_row":
        apply_row_based_single_hue_fill_to_excel(
            worksheet,
            pivot_df,
            start_row,
            start_column,
            end_color="#E06666",
            include_totals=False,
        )
        return
    if color_mode == "weapon_evacuation_row":
        apply_row_based_centered_fill_to_excel(
            worksheet,
            pivot_df,
            start_row,
            start_column,
            positive_end_color="#E06666",
            negative_end_color="#93C47D",
            include_totals=False,
        )
        apply_centered_summary_fill_to_excel(
            worksheet,
            pivot_df,
            start_row,
            start_column,
            positive_end_color="#E06666",
            negative_end_color="#93C47D",
        )
        return
    if color_mode == "delta_diverging":
        apply_diverging_fill_to_excel(
            worksheet,
            pivot_df,
            start_row,
            start_column,
            positive_end_color="#E06666",
            negative_end_color="#93C47D",
            include_totals=True,
        )
        return

    apply_single_hue_fill_to_excel(
        worksheet,
        pivot_df,
        start_row,
        start_column,
        end_color=f"#{block_definition['end_color']}",
        include_totals=False,
    )


def apply_feishu_single_hue_fill_to_block(
    sheet_id: str,
    pivot_df: pd.DataFrame,
    start_row: int,
    start_column: int,
    end_color: str,
    include_totals: bool = False,
) -> list[dict[str, Any]]:
    body_rows, body_columns = iter_matrix_labels(pivot_df, include_totals=include_totals)
    style_requests: list[dict[str, Any]] = []

    numeric_values: list[float] = []
    for row_name in body_rows:
        for column_name in body_columns:
            value = pivot_df.loc[row_name, column_name]
            if pd.isna(value):
                continue
            numeric_values.append(float(value))

    if not numeric_values:
        return style_requests

    min_value = min(numeric_values)
    max_value = max(numeric_values)

    for row_offset, row_name in enumerate(body_rows):
        excel_row = start_row + 2 + row_offset
        for column_offset, column_name in enumerate(body_columns, start=start_column + 1):
            value = pivot_df.loc[row_name, column_name]
            if pd.isna(value):
                continue

            numeric_value = float(value)
            if max_value == min_value:
                ratio = 1.0 if numeric_value > 0 else 0.0
            else:
                ratio = (numeric_value - min_value) / (max_value - min_value)

            if ratio <= 0:
                continue

            background_color = interpolate_hex_color("#FFFFFF", end_color, ratio)
            column_letter = build_excel_column_name(column_offset)
            cell_range = f"{sheet_id}!{column_letter}{excel_row}:{column_letter}{excel_row}"
            style_requests.append(
                build_background_style_request([cell_range], background_color)
            )

    return style_requests


def apply_feishu_row_based_single_hue_fill_to_block(
    sheet_id: str,
    pivot_df: pd.DataFrame,
    start_row: int,
    start_column: int,
    end_color: str,
    include_totals: bool = False,
) -> list[dict[str, Any]]:
    row_labels, column_labels = iter_matrix_labels(pivot_df, include_totals=include_totals)
    style_requests: list[dict[str, Any]] = []

    for row_offset, row_name in enumerate(row_labels):
        numeric_values = [
            float(pivot_df.loc[row_name, column_name])
            for column_name in column_labels
            if not pd.isna(pivot_df.loc[row_name, column_name])
        ]
        if not numeric_values:
            continue

        row_min = min(numeric_values)
        row_max = max(numeric_values)
        excel_row = start_row + 2 + row_offset

        for column_offset, column_name in enumerate(column_labels, start=start_column + 1):
            value = pivot_df.loc[row_name, column_name]
            if pd.isna(value):
                continue

            numeric_value = float(value)
            if row_max == row_min:
                ratio = 1.0 if numeric_value > 0 else 0.0
            else:
                ratio = (numeric_value - row_min) / (row_max - row_min)

            if ratio <= 0:
                continue

            column_letter = build_excel_column_name(column_offset)
            cell_range = f"{sheet_id}!{column_letter}{excel_row}:{column_letter}{excel_row}"
            style_requests.append(
                build_background_style_request(
                    [cell_range],
                    interpolate_hex_color("#FFFFFF", end_color, ratio),
                )
            )

    return style_requests


def apply_feishu_row_based_centered_fill_to_block(
    sheet_id: str,
    pivot_df: pd.DataFrame,
    start_row: int,
    start_column: int,
    positive_end_color: str,
    negative_end_color: str,
    include_totals: bool = False,
) -> list[dict[str, Any]]:
    row_labels, column_labels = iter_matrix_labels(pivot_df, include_totals=include_totals)
    style_requests: list[dict[str, Any]] = []

    for row_offset, row_name in enumerate(row_labels):
        if "*" in pivot_df.columns and not pd.isna(pivot_df.loc[row_name, "*"]):
            center_value = float(pivot_df.loc[row_name, "*"])
        else:
            center_value = None

        numeric_values = [
            float(pivot_df.loc[row_name, column_name])
            for column_name in column_labels
            if not pd.isna(pivot_df.loc[row_name, column_name])
        ]
        if not numeric_values:
            continue
        if center_value is None:
            center_value = sum(numeric_values) / len(numeric_values)

        row_min = min(numeric_values)
        row_max = max(numeric_values)
        excel_row = start_row + 2 + row_offset

        for column_offset, column_name in enumerate(column_labels, start=start_column + 1):
            value = pivot_df.loc[row_name, column_name]
            if pd.isna(value):
                continue

            numeric_value = float(value)
            if numeric_value >= center_value:
                denominator = max(row_max - center_value, 0.0)
                ratio = 0.0 if denominator == 0 else (numeric_value - center_value) / denominator
                target_color = positive_end_color
            else:
                denominator = max(center_value - row_min, 0.0)
                ratio = 0.0 if denominator == 0 else (center_value - numeric_value) / denominator
                target_color = negative_end_color

            if ratio <= 0:
                continue

            column_letter = build_excel_column_name(column_offset)
            cell_range = f"{sheet_id}!{column_letter}{excel_row}:{column_letter}{excel_row}"
            style_requests.append(
                build_background_style_request(
                    [cell_range],
                    interpolate_hex_color("#FFFFFF", target_color, ratio),
                )
            )

    return style_requests


def apply_feishu_diverging_fill_to_block(
    sheet_id: str,
    pivot_df: pd.DataFrame,
    start_row: int,
    start_column: int,
    positive_end_color: str,
    negative_end_color: str,
    include_totals: bool = True,
) -> list[dict[str, Any]]:
    row_labels, column_labels = iter_matrix_labels(pivot_df, include_totals=include_totals)
    style_requests: list[dict[str, Any]] = []
    positive_max = 0.0
    negative_max = 0.0

    for row_name in row_labels:
        for column_name in column_labels:
            value = pivot_df.loc[row_name, column_name]
            if pd.isna(value):
                continue
            numeric_value = float(value)
            if numeric_value > 0:
                positive_max = max(positive_max, numeric_value)
            elif numeric_value < 0:
                negative_max = max(negative_max, abs(numeric_value))

    if positive_max == 0 and negative_max == 0:
        return style_requests

    for row_offset, row_name in enumerate(row_labels):
        excel_row = start_row + 2 + row_offset
        for column_offset, column_name in enumerate(column_labels, start=start_column + 1):
            value = pivot_df.loc[row_name, column_name]
            if pd.isna(value):
                continue

            numeric_value = float(value)
            if numeric_value > 0 and positive_max > 0:
                ratio = numeric_value / positive_max
                target_color = positive_end_color
            elif numeric_value < 0 and negative_max > 0:
                ratio = abs(numeric_value) / negative_max
                target_color = negative_end_color
            else:
                continue

            column_letter = build_excel_column_name(column_offset)
            cell_range = f"{sheet_id}!{column_letter}{excel_row}:{column_letter}{excel_row}"
            style_requests.append(
                build_background_style_request(
                    [cell_range],
                    interpolate_hex_color("#FFFFFF", target_color, ratio),
                )
            )

    return style_requests


def apply_feishu_centered_summary_fill_to_block(
    sheet_id: str,
    pivot_df: pd.DataFrame,
    start_row: int,
    start_column: int,
    positive_end_color: str,
    negative_end_color: str,
) -> list[dict[str, Any]]:
    if "*" not in pivot_df.index or "*" not in pivot_df.columns:
        return []

    grand_total = pivot_df.loc["*", "*"]
    if pd.isna(grand_total):
        return []
    center_value = float(grand_total)
    style_requests: list[dict[str, Any]] = []

    total_column_index = pivot_df.columns.get_loc("*") + start_column + 1
    total_column_letter = build_excel_column_name(total_column_index)
    total_row_index = start_row + 2 + pivot_df.index.get_loc("*")

    hero_summary_values = [
        float(pivot_df.loc[row_name, "*"])
        for row_name in pivot_df.index
        if row_name != "*" and not pd.isna(pivot_df.loc[row_name, "*"])
    ]
    weapon_summary_values = [
        float(pivot_df.loc["*", column_name])
        for column_name in pivot_df.columns
        if column_name != "*" and not pd.isna(pivot_df.loc["*", column_name])
    ]

    hero_min = min(hero_summary_values) if hero_summary_values else center_value
    hero_max = max(hero_summary_values) if hero_summary_values else center_value
    weapon_min = min(weapon_summary_values) if weapon_summary_values else center_value
    weapon_max = max(weapon_summary_values) if weapon_summary_values else center_value

    for row_offset, row_name in enumerate(pivot_df.index):
        if row_name == "*":
            continue
        value = pivot_df.loc[row_name, "*"]
        if pd.isna(value):
            continue
        numeric_value = float(value)
        if numeric_value >= center_value:
            denominator = max(hero_max - center_value, 0.0)
            ratio = 0.0 if denominator == 0 else (numeric_value - center_value) / denominator
            target_color = positive_end_color
        else:
            denominator = max(center_value - hero_min, 0.0)
            ratio = 0.0 if denominator == 0 else (center_value - numeric_value) / denominator
            target_color = negative_end_color
        if ratio <= 0:
            continue
        excel_row = start_row + 2 + row_offset
        style_requests.append(
            build_background_style_request(
                [f"{sheet_id}!{total_column_letter}{excel_row}:{total_column_letter}{excel_row}"],
                interpolate_hex_color("#FFFFFF", target_color, ratio),
            )
        )

    for column_offset, column_name in enumerate(pivot_df.columns, start=start_column + 1):
        if column_name == "*":
            continue
        value = pivot_df.loc["*", column_name]
        if pd.isna(value):
            continue
        numeric_value = float(value)
        if numeric_value >= center_value:
            denominator = max(weapon_max - center_value, 0.0)
            ratio = 0.0 if denominator == 0 else (numeric_value - center_value) / denominator
            target_color = positive_end_color
        else:
            denominator = max(center_value - weapon_min, 0.0)
            ratio = 0.0 if denominator == 0 else (center_value - numeric_value) / denominator
            target_color = negative_end_color
        if ratio <= 0:
            continue
        column_letter = build_excel_column_name(column_offset)
        style_requests.append(
            build_background_style_request(
                [f"{sheet_id}!{column_letter}{total_row_index}:{column_letter}{total_row_index}"],
                interpolate_hex_color("#FFFFFF", target_color, ratio),
            )
        )

    return style_requests


def apply_feishu_fill_to_block(
    sheet_id: str,
    pivot_df: pd.DataFrame,
    start_row: int,
    start_column: int,
    block_definition: dict[str, Any],
) -> list[dict[str, Any]]:
    color_mode = block_definition.get("color_mode", "global_single")
    if color_mode == "weapon_usage_row":
        return apply_feishu_row_based_single_hue_fill_to_block(
            sheet_id,
            pivot_df,
            start_row,
            start_column,
            end_color="#E06666",
            include_totals=False,
        )
    if color_mode == "weapon_evacuation_row":
        return (
            apply_feishu_row_based_centered_fill_to_block(
                sheet_id,
                pivot_df,
                start_row,
                start_column,
                positive_end_color="#E06666",
                negative_end_color="#93C47D",
                include_totals=False,
            )
            + apply_feishu_centered_summary_fill_to_block(
                sheet_id,
                pivot_df,
                start_row,
                start_column,
                positive_end_color="#E06666",
                negative_end_color="#93C47D",
            )
        )
    if color_mode == "delta_diverging":
        return apply_feishu_diverging_fill_to_block(
            sheet_id,
            pivot_df,
            start_row,
            start_column,
            positive_end_color="#E06666",
            negative_end_color="#93C47D",
            include_totals=True,
        )

    return apply_feishu_single_hue_fill_to_block(
        sheet_id,
        pivot_df,
        start_row,
        start_column,
        end_color=f"#{block_definition['end_color']}",
        include_totals=False,
    )


def build_usage_rate_summary_style_requests(
    sheet_id: str,
    usage_rate_df: pd.DataFrame,
    start_row: int,
    start_column: int,
    hero_low_threshold: float,
    hero_high_threshold: float,
    weapon_low_threshold: float,
    weapon_high_threshold: float,
) -> list[dict[str, Any]]:
    if "*" not in usage_rate_df.index or "*" not in usage_rate_df.columns:
        return []

    style_requests: list[dict[str, Any]] = []
    total_column_index = usage_rate_df.columns.get_loc("*") + start_column + 1
    total_column_letter = build_excel_column_name(total_column_index)
    total_row_index = start_row + 2 + usage_rate_df.index.get_loc("*")

    for row_offset, row_name in enumerate(usage_rate_df.index):
        if row_name == "*":
            continue

        value = usage_rate_df.loc[row_name, "*"]
        if pd.isna(value):
            continue

        fill_color = ""
        numeric_value = float(value)
        if numeric_value < hero_low_threshold:
            fill_color = SUMMARY_LOW_FILL_COLOR
        elif numeric_value > hero_high_threshold:
            fill_color = SUMMARY_HIGH_FILL_COLOR
        if not fill_color:
            continue
        excel_row = start_row + 2 + row_offset
        style_requests.append(
            build_background_style_request(
                [f"{sheet_id}!{total_column_letter}{excel_row}:{total_column_letter}{excel_row}"],
                fill_color,
            )
        )

    for column_offset, column_name in enumerate(usage_rate_df.columns, start=start_column + 1):
        if column_name == "*":
            continue

        value = usage_rate_df.loc["*", column_name]
        if pd.isna(value):
            continue

        fill_color = ""
        numeric_value = float(value)
        if numeric_value < weapon_low_threshold:
            fill_color = SUMMARY_LOW_FILL_COLOR
        elif numeric_value > weapon_high_threshold:
            fill_color = SUMMARY_HIGH_FILL_COLOR
        if not fill_color:
            continue
        column_letter = build_excel_column_name(column_offset)
        style_requests.append(
            build_background_style_request(
                [f"{sheet_id}!{column_letter}{total_row_index}:{column_letter}{total_row_index}"],
                fill_color,
            )
        )

    return style_requests


def normalize_lookup_key(value: Any) -> str:
    if pd.isna(value):
        return ""

    if isinstance(value, int):
        return str(value)

    if isinstance(value, float) and value.is_integer():
        return str(int(value))

    text = str(value).strip()
    if not text:
        return ""

    try:
        float_value = float(text)
    except ValueError:
        return text

    if float_value.is_integer():
        return str(int(float_value))

    return text


def load_item_name_mapping(mapping_excel_path: Path) -> dict[str, str]:
    if not mapping_excel_path.exists():
        raise FileNotFoundError(f"未找到映射文件：{mapping_excel_path}")

    excel_file = pd.ExcelFile(mapping_excel_path, engine="openpyxl")
    id_name_mapping: dict[str, str] = {}
    name_column_candidates = ("道具名", "名称", "中文名", "英雄名", "武器名", "装备名")

    for sheet_name in excel_file.sheet_names:
        sheet_df = pd.read_excel(excel_file, sheet_name=sheet_name, engine="openpyxl")
        if sheet_df.empty:
            continue

        normalized_columns = {str(column).strip(): column for column in sheet_df.columns}
        id_column = normalized_columns.get("ID") or normalized_columns.get("id")
        name_column = next(
            (
                normalized_columns[column_name]
                for column_name in name_column_candidates
                if column_name in normalized_columns
            ),
            None,
        )

        if id_column is None or name_column is None:
            continue

        valid_rows_df = sheet_df[[id_column, name_column]].dropna(subset=[id_column, name_column])
        for item_id, item_name in valid_rows_df.itertuples(index=False):
            lookup_key = normalize_lookup_key(item_id)
            display_name = str(item_name).strip()
            if lookup_key and display_name:
                id_name_mapping[lookup_key] = display_name

    if not id_name_mapping:
        raise ValueError(f"映射文件中未读取到有效的 ID 和中文名称：{mapping_excel_path}")

    return id_name_mapping


def load_equipment_set_reference(mapping_excel_path: Path) -> tuple[dict[str, str], list[str]]:
    if not mapping_excel_path.exists():
        raise FileNotFoundError(f"未找到映射文件：{mapping_excel_path}")

    excel_file = pd.ExcelFile(mapping_excel_path, engine="openpyxl")
    equipment_name_to_set_mapping: dict[str, str] = {}
    set_name_pool: set[str] = set()

    for sheet_name in excel_file.sheet_names:
        sheet_df = pd.read_excel(excel_file, sheet_name=sheet_name, engine="openpyxl")
        if sheet_df.empty:
            continue

        normalized_columns = {str(column).strip(): column for column in sheet_df.columns}
        name_column = (
            normalized_columns.get("道具名")
            or normalized_columns.get("名称")
            or normalized_columns.get("装备名")
        )
        set_column = normalized_columns.get("套装")

        if name_column is None or set_column is None:
            continue

        valid_rows_df = sheet_df[[name_column, set_column]].dropna(subset=[name_column, set_column])
        for item_name, set_name in valid_rows_df.itertuples(index=False):
            display_name = str(item_name).strip()
            normalized_set_name = str(set_name).strip()
            if not display_name or not normalized_set_name:
                continue

            equipment_name_to_set_mapping[display_name] = normalized_set_name
            set_name_pool.add(normalized_set_name)

    return equipment_name_to_set_mapping, sorted(set_name_pool, key=lambda value: (-len(value), value))


def translate_single_value(value: Any, id_name_mapping: dict[str, str]) -> Any:
    if pd.isna(value):
        return value

    lookup_key = normalize_lookup_key(value)
    if not lookup_key:
        return value

    return id_name_mapping.get(lookup_key, value)


def translate_result_dataframe(
    result_df: pd.DataFrame,
    id_name_mapping: dict[str, str],
) -> pd.DataFrame:
    translated_df = result_df.copy()

    for column_name in TRANSLATE_COLUMNS:
        if column_name in translated_df.columns:
            source_series = translated_df[column_name]
            unique_values = pd.unique(source_series.dropna())
            translation_mapping = {
                value: translate_single_value(value, id_name_mapping)
                for value in unique_values
            }
            translated_series = source_series.map(translation_mapping)
            translated_df[column_name] = translated_series.where(
                translated_series.notna(),
                source_series,
            )

    return translated_df


def _extract_weapon_level(weapon_type_value: Any) -> int:
    weapon_type_text = str(weapon_type_value).strip()
    if not weapon_type_text or weapon_type_text == "0":
        return -1

    last_character = weapon_type_text[-1]
    if last_character.isdigit():
        return int(last_character)

    return -1


def normalize_weapon_source_columns(result_df: pd.DataFrame) -> pd.DataFrame:
    normalized_df = result_df.copy()
    required_columns = (
        "b_main_weapon_type",
        "b_main_weapon_variant",
        "b_main_weapon_ult",
        "b_main_weapon_boost",
        "b_sub_weapon_type",
        "b_sub_weapon_variant",
        "b_sub_weapon_ult",
        "b_sub_weapon_boost",
    )
    if any(column_name not in normalized_df.columns for column_name in required_columns):
        return normalized_df

    main_weapon_type_series = normalized_df["b_main_weapon_type"].fillna("").astype(str).str.strip()
    sub_weapon_type_series = normalized_df["b_sub_weapon_type"].fillna("").astype(str).str.strip()
    both_weapon_present_mask = (
        main_weapon_type_series.ne("")
        & main_weapon_type_series.ne("0")
        & sub_weapon_type_series.ne("")
        & sub_weapon_type_series.ne("0")
    )
    if not both_weapon_present_mask.any():
        return normalized_df

    main_weapon_level_series = main_weapon_type_series.map(_extract_weapon_level)
    sub_weapon_level_series = sub_weapon_type_series.map(_extract_weapon_level)
    use_sub_weapon_mask = both_weapon_present_mask & sub_weapon_level_series.gt(main_weapon_level_series)

    if not use_sub_weapon_mask.any():
        return normalized_df

    for suffix in ("type", "variant", "ult", "boost"):
        normalized_df.loc[use_sub_weapon_mask, f"b_main_weapon_{suffix}"] = normalized_df.loc[
            use_sub_weapon_mask, f"b_sub_weapon_{suffix}"
        ]

    normalized_df["b_main_weapon_level"] = normalized_df["b_main_weapon_type"].map(_extract_weapon_level)
    return normalized_df


def normalize_equipment_series(equipment_series: pd.Series) -> pd.Series:
    normalized_series = equipment_series.fillna("").astype(str).str.strip()
    normalized_series = normalized_series.mask(normalized_series == "0", "")
    normalized_series = normalized_series.str.replace(r"^\[[^\]]+\]", "", regex=True).str.strip()

    prefix_pattern = "^(" + "|".join(re.escape(prefix) for prefix in EQUIPMENT_NAME_PREFIXES) + ")"
    normalized_series = normalized_series.str.replace(prefix_pattern, "", regex=True).str.strip()
    return normalized_series


def infer_equipment_set_series(
    equipment_series: pd.Series,
    equipment_name_to_set_mapping: dict[str, str],
    set_name_candidates: list[str],
) -> pd.Series:
    normalized_series = normalize_equipment_series(equipment_series)
    set_series = normalized_series.map(equipment_name_to_set_mapping).fillna("")

    for set_name in set_name_candidates:
        missing_mask = set_series.eq("") & normalized_series.str.contains(re.escape(set_name), regex=True)
        set_series = set_series.mask(missing_mask, set_name)

    return set_series


def build_equipment_set_series(
    pivot_source_df: pd.DataFrame,
    equipment_name_to_set_mapping: dict[str, str],
    set_name_candidates: list[str],
) -> pd.Series:
    headset_set_series = infer_equipment_set_series(
        pivot_source_df["b_headset"],
        equipment_name_to_set_mapping,
        set_name_candidates,
    )
    armor_set_series = infer_equipment_set_series(
        pivot_source_df["b_armor"],
        equipment_name_to_set_mapping,
        set_name_candidates,
    )
    shoe_set_series = infer_equipment_set_series(
        pivot_source_df["b_shoe"],
        equipment_name_to_set_mapping,
        set_name_candidates,
    )

    same_set_mask = (
        headset_set_series.ne("")
        & headset_set_series.eq(armor_set_series)
        & headset_set_series.eq(shoe_set_series)
    )
    equipment_set_series = pd.Series("散件", index=pivot_source_df.index, dtype="object")
    equipment_set_series.loc[same_set_mask] = (
        headset_set_series.loc[same_set_mask] + "套装"
    )
    return equipment_set_series


def prepare_pivot_source_dataframe(
    result_df: pd.DataFrame,
    equipment_name_to_set_mapping: dict[str, str],
    set_name_candidates: list[str],
) -> pd.DataFrame:
    # 透视前先统一清洗字段，并拼出“套装”和“攻击部”字段供后续透视直接复用。
    pivot_source_df = result_df.loc[
        :,
        [
            "b_battle_begin_hero",
            "b_main_weapon_type",
            "b_main_weapon_variant",
            "b_main_weapon_ult",
            "b_main_weapon_boost",
            "b_main_weapon_level",
            "b_headset",
            "b_armor",
            "b_shoe",
            "a_evacuatedsucc",
            "a_battleid",
            "a_battletime",
            "a_inc_value",
            "a_moved",
            "a_player_down",
            "a_monster_killed",
        ],
    ].copy()
    pivot_source_df["b_battle_begin_hero"] = (
        pivot_source_df["b_battle_begin_hero"].fillna("").astype(str).str.strip()
    )
    pivot_source_df["b_main_weapon_type"] = (
        pivot_source_df["b_main_weapon_type"].fillna("").astype(str).str.strip()
    )
    pivot_source_df["b_main_weapon_variant"] = (
        pivot_source_df["b_main_weapon_variant"].fillna("").astype(str).str.strip()
    )
    pivot_source_df["b_main_weapon_ult"] = (
        pivot_source_df["b_main_weapon_ult"].fillna("").astype(str).str.strip()
    )
    pivot_source_df["b_main_weapon_boost"] = (
        pivot_source_df["b_main_weapon_boost"].fillna("").astype(str).str.strip()
    )
    pivot_source_df["b_main_weapon_level"] = pd.to_numeric(
        pivot_source_df["b_main_weapon_level"],
        errors="coerce",
    )
    for column_name in (
        "b_main_weapon_type",
        "b_main_weapon_variant",
        "b_main_weapon_ult",
        "b_main_weapon_boost",
        "b_headset",
        "b_armor",
        "b_shoe",
    ):
        pivot_source_df[column_name] = pivot_source_df[column_name].fillna("").astype(str).str.strip()
        pivot_source_df.loc[pivot_source_df[column_name] == "0", column_name] = ""
    pivot_source_df["b_equipment_set"] = build_equipment_set_series(
        pivot_source_df,
        equipment_name_to_set_mapping=equipment_name_to_set_mapping,
        set_name_candidates=set_name_candidates,
    )
    pivot_source_df["b_main_weapon"] = pivot_source_df["b_main_weapon_type"]
    variant_mask = pivot_source_df["b_main_weapon_variant"].ne("")
    pivot_source_df.loc[variant_mask, "b_main_weapon"] = (
        pivot_source_df.loc[variant_mask, "b_main_weapon_type"]
        + "-"
        + pivot_source_df.loc[variant_mask, "b_main_weapon_variant"]
    )
    pivot_source_df["b_main_weapon_with_ult"] = pivot_source_df["b_main_weapon_type"]
    ult_mask = pivot_source_df["b_main_weapon_ult"].ne("")
    pivot_source_df.loc[ult_mask, "b_main_weapon_with_ult"] = (
        pivot_source_df.loc[ult_mask, "b_main_weapon_type"]
        + "-"
        + pivot_source_df.loc[ult_mask, "b_main_weapon_ult"]
    )
    pivot_source_df["a_evacuatedsucc"] = pd.to_numeric(
        pivot_source_df["a_evacuatedsucc"],
        errors="coerce",
    )
    pivot_source_df["a_inc_value"] = pd.to_numeric(
        pivot_source_df["a_inc_value"],
        errors="coerce",
    )
    pivot_source_df["a_battletime"] = pd.to_numeric(
        pivot_source_df["a_battletime"],
        errors="coerce",
    )
    pivot_source_df["a_moved"] = pd.to_numeric(
        pivot_source_df["a_moved"],
        errors="coerce",
    )
    pivot_source_df["a_player_down"] = pd.to_numeric(
        pivot_source_df["a_player_down"],
        errors="coerce",
    )
    pivot_source_df["a_monster_killed"] = pd.to_numeric(
        pivot_source_df["a_monster_killed"],
        errors="coerce",
    )
    valid_battletime_mask = pivot_source_df["a_battletime"] > 0
    pivot_source_df["a_moved_per_minute"] = float("nan")
    pivot_source_df.loc[valid_battletime_mask, "a_moved_per_minute"] = (
        pivot_source_df.loc[valid_battletime_mask, "a_moved"] * 60
        / pivot_source_df.loc[valid_battletime_mask, "a_battletime"]
    )

    pivot_source_df = pivot_source_df.loc[
        (pivot_source_df["b_battle_begin_hero"] != "")
        & (pivot_source_df["b_battle_begin_hero"] != "0")
        & (pivot_source_df["b_main_weapon_type"] != "")
    ].copy()

    return pivot_source_df


def build_weapon_sort_key(weapon_label: str) -> tuple[int, str]:
    weapon_type = weapon_label.split("-", 1)[0].strip()
    return (
        MAIN_WEAPON_TYPE_ORDER_MAP.get(weapon_type, len(MAIN_WEAPON_TYPE_ORDER)),
        weapon_label,
    )


def build_hero_weapon_pivot_tables(
    pivot_source_df: pd.DataFrame,
    weapon_column_name: str,
    column_sort_key: Callable[[str], tuple[Any, ...]] | None = None,
) -> dict[str, pd.DataFrame]:
    pivot_source_df = pivot_source_df.loc[pivot_source_df[weapon_column_name] != ""].copy()

    if pivot_source_df.empty:
        empty_df = pd.DataFrame(index=["*"], columns=["*"], data=[[0]])
        return {
            "usage_rate_df": empty_df.astype(float),
            "evacuation_rate_df": empty_df.astype(float),
            "avg_inc_value_df": empty_df.astype(float),
            "avg_moved_per_minute_df": empty_df.astype(float),
            "avg_player_down_df": empty_df.astype(float),
            "avg_monster_killed_df": empty_df.astype(float),
            "count_df": empty_df.astype(int),
        }

    metric_column_mapping = {
        "evacuation_rate": "a_evacuatedsucc",
        "avg_inc_value": "a_inc_value",
        "avg_moved_per_minute": "a_moved_per_minute",
        "avg_player_down": "a_player_down",
        "avg_monster_killed": "a_monster_killed",
    }
    aggregate_kwargs: dict[str, tuple[str, str]] = {
        "count": ("a_battleid", "size"),
    }
    for metric_name, source_column_name in metric_column_mapping.items():
        aggregate_kwargs[f"{metric_name}_sum"] = (source_column_name, "sum")
        aggregate_kwargs[f"{metric_name}_count"] = (source_column_name, "count")

    grouped_df = pivot_source_df.groupby(
        ["b_battle_begin_hero", weapon_column_name],
        sort=False,
        dropna=False,
    ).agg(**aggregate_kwargs)

    def build_metric_series(summary_df: pd.DataFrame, metric_name: str) -> pd.Series:
        if metric_name == "count":
            return summary_df["count"]
        denominator = summary_df[f"{metric_name}_count"].replace(0, float("nan"))
        return summary_df[f"{metric_name}_sum"] / denominator

    def build_metric_matrix(metric_name: str, fill_value: float | int | None = None) -> pd.DataFrame:
        metric_df = build_metric_series(grouped_df, metric_name).unstack(weapon_column_name)
        if fill_value is not None:
            metric_df = metric_df.fillna(fill_value)
        return metric_df

    count_df = build_metric_matrix("count", fill_value=0)
    evacuation_rate_df = build_metric_matrix("evacuation_rate")
    avg_inc_value_df = build_metric_matrix("avg_inc_value")
    avg_moved_per_minute_df = build_metric_matrix("avg_moved_per_minute")
    avg_player_down_df = build_metric_matrix("avg_player_down")
    avg_monster_killed_df = build_metric_matrix("avg_monster_killed")

    row_totals = grouped_df.groupby(level=0, sort=False).sum()
    column_totals = grouped_df.groupby(level=1, sort=False).sum()
    grand_totals = grouped_df.sum()

    def append_margins(metric_df: pd.DataFrame, metric_name: str, fill_value: float | int | None = None) -> pd.DataFrame:
        metric_with_margins = metric_df.copy()
        metric_with_margins["*"] = build_metric_series(row_totals, metric_name)
        total_row = build_metric_series(column_totals, metric_name).to_dict()
        total_row["*"] = build_metric_series(grand_totals.to_frame().T, metric_name).iloc[0]
        metric_with_margins.loc["*"] = total_row
        if fill_value is not None:
            metric_with_margins = metric_with_margins.fillna(fill_value)
        return metric_with_margins

    count_df = append_margins(count_df, "count", fill_value=0)
    evacuation_rate_df = append_margins(evacuation_rate_df, "evacuation_rate")
    avg_inc_value_df = append_margins(avg_inc_value_df, "avg_inc_value")
    avg_moved_per_minute_df = append_margins(avg_moved_per_minute_df, "avg_moved_per_minute")
    avg_player_down_df = append_margins(avg_player_down_df, "avg_player_down")
    avg_monster_killed_df = append_margins(avg_monster_killed_df, "avg_monster_killed")

    hero_labels = [label for label in count_df.index.tolist() if label != "*"]
    weapon_labels = [label for label in count_df.columns.tolist() if label != "*"]
    hero_labels.sort(
        key=lambda label: (
            -int(count_df.loc[label, weapon_labels].sum()),
            label,
        )
    )
    if column_sort_key is None:
        weapon_labels.sort(
            key=lambda label: (
                -int(count_df.loc["*", label]),
                label,
            )
        )
    else:
        weapon_labels.sort(key=column_sort_key)

    ordered_rows = hero_labels + ["*"]
    ordered_columns = weapon_labels + ["*"]
    count_df = count_df.reindex(index=ordered_rows, columns=ordered_columns, fill_value=0)
    evacuation_rate_df = evacuation_rate_df.reindex(index=ordered_rows, columns=ordered_columns)
    avg_inc_value_df = avg_inc_value_df.reindex(index=ordered_rows, columns=ordered_columns)
    avg_moved_per_minute_df = avg_moved_per_minute_df.reindex(index=ordered_rows, columns=ordered_columns)
    avg_player_down_df = avg_player_down_df.reindex(index=ordered_rows, columns=ordered_columns)
    avg_monster_killed_df = avg_monster_killed_df.reindex(index=ordered_rows, columns=ordered_columns)

    if hero_labels and weapon_labels:
        valid_mask = count_df.loc[hero_labels, weapon_labels] > 0
        evacuation_rate_df.loc[hero_labels, weapon_labels] = evacuation_rate_df.loc[
            hero_labels, weapon_labels
        ].where(valid_mask)
        avg_inc_value_df.loc[hero_labels, weapon_labels] = avg_inc_value_df.loc[
            hero_labels, weapon_labels
        ].where(valid_mask)
        avg_moved_per_minute_df.loc[hero_labels, weapon_labels] = avg_moved_per_minute_df.loc[
            hero_labels, weapon_labels
        ].where(valid_mask)
        avg_player_down_df.loc[hero_labels, weapon_labels] = avg_player_down_df.loc[
            hero_labels, weapon_labels
        ].where(valid_mask)
        avg_monster_killed_df.loc[hero_labels, weapon_labels] = avg_monster_killed_df.loc[
            hero_labels, weapon_labels
        ].where(valid_mask)

    count_df = count_df.astype(int)
    total_count = int(count_df.loc["*", "*"])
    if total_count == 0:
        usage_rate_df = count_df.astype(float)
    else:
        usage_rate_df = count_df.astype(float) / total_count

    return {
        "usage_rate_df": usage_rate_df,
        "evacuation_rate_df": evacuation_rate_df,
        "avg_inc_value_df": avg_inc_value_df,
        "avg_moved_per_minute_df": avg_moved_per_minute_df,
        "avg_player_down_df": avg_player_down_df,
        "avg_monster_killed_df": avg_monster_killed_df,
        "count_df": count_df,
    }


def style_summary_sheet(worksheet: openpyxl.worksheet.worksheet.Worksheet) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="E7E6E6")
    title_fill = PatternFill(fill_type="solid", fgColor="D9D9D9")
    centered_alignment = Alignment(horizontal="center", vertical="center")
    bold_font = Font(bold=True)
    normal_font = Font(bold=False)
    existing_cells = list(worksheet._cells.values())
    if not existing_cells:
        return

    max_row = max(cell.row for cell in existing_cells)
    cells_by_row: dict[int, list[openpyxl.cell.cell.Cell]] = {}
    max_lengths_by_column: dict[int, int] = {}

    for cell in existing_cells:
        cell.alignment = centered_alignment
        cell.font = bold_font if should_bold_label(cell.value) else normal_font
        cells_by_row.setdefault(cell.row, []).append(cell)
        cell_value = "" if cell.value is None else str(cell.value)
        max_lengths_by_column[cell.column] = max(
            max_lengths_by_column.get(cell.column, 0),
            len(cell_value),
        )

    for cell in cells_by_row.get(1, []):
        if cell.value is not None:
            cell.fill = title_fill

    for cell in cells_by_row.get(2, []):
        if cell.value is not None:
            cell.fill = header_fill

    for cell in cells_by_row.get(max_row, []):
        if cell.value is not None:
            cell.fill = header_fill

    for row_index in range(3, max_row + 1):
        row_cells = cells_by_row.get(row_index, [])
        title_cell = next((cell for cell in row_cells if cell.column == 1), None)
        title_value = title_cell.value if title_cell is not None else None
        if title_value in ALL_BLOCK_TITLES:
            for cell in row_cells:
                if cell.value is not None:
                    cell.fill = title_fill
            for cell in cells_by_row.get(row_index + 1, []):
                if cell.value is not None:
                    cell.fill = header_fill

    for column_index, max_length in max_lengths_by_column.items():
        column_letter = build_excel_column_name(column_index)
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 4, 10), 18)


def write_pivot_block(
    worksheet: openpyxl.worksheet.worksheet.Worksheet,
    start_row: int,
    start_column: int,
    title: str,
    pivot_df: pd.DataFrame,
    percentage: bool = False,
) -> tuple[int, int, int]:
    worksheet.cell(row=start_row, column=start_column, value=title)

    header_row = start_row + 1
    worksheet.cell(row=header_row, column=start_column, value="")
    for column_offset, column_name in enumerate(pivot_df.columns, start=start_column + 1):
        worksheet.cell(row=header_row, column=column_offset, value=column_name)

    for row_offset, row_name in enumerate(pivot_df.index, start=1):
        worksheet.cell(row=header_row + row_offset, column=start_column, value=row_name)
        for column_offset, column_name in enumerate(pivot_df.columns, start=start_column + 1):
            value = pivot_df.loc[row_name, column_name]
            cell = worksheet.cell(
                row=header_row + row_offset,
                column=column_offset,
                value=None if pd.isna(value) else value,
            )
            if percentage and value is not None and not pd.isna(value):
                cell.number_format = "0%"

    return header_row, header_row + len(pivot_df.index), start_column + len(pivot_df.columns)


def apply_usage_rate_summary_fill_to_excel(
    worksheet: openpyxl.worksheet.worksheet.Worksheet,
    usage_rate_df: pd.DataFrame,
    start_row: int,
    start_column: int,
    hero_low_threshold: float,
    hero_high_threshold: float,
    weapon_low_threshold: float,
    weapon_high_threshold: float,
) -> None:
    if "*" not in usage_rate_df.index or "*" not in usage_rate_df.columns:
        return

    total_column_index = usage_rate_df.columns.get_loc("*") + start_column + 1
    total_row_index = start_row + 2 + usage_rate_df.index.get_loc("*")

    for row_offset, row_name in enumerate(usage_rate_df.index):
        if row_name == "*":
            continue

        value = usage_rate_df.loc[row_name, "*"]
        if pd.isna(value):
            continue

        fill_color = ""
        numeric_value = float(value)
        if numeric_value < hero_low_threshold:
            fill_color = SUMMARY_LOW_FILL_COLOR
        elif numeric_value > hero_high_threshold:
            fill_color = SUMMARY_HIGH_FILL_COLOR
        if not fill_color:
            continue
        excel_row = start_row + 2 + row_offset
        worksheet.cell(row=excel_row, column=total_column_index).fill = PatternFill(
            fill_type="solid",
            fgColor=fill_color.lstrip("#"),
        )

    for column_offset, column_name in enumerate(usage_rate_df.columns, start=start_column + 1):
        if column_name == "*":
            continue

        value = usage_rate_df.loc["*", column_name]
        if pd.isna(value):
            continue

        fill_color = ""
        numeric_value = float(value)
        if numeric_value < weapon_low_threshold:
            fill_color = SUMMARY_LOW_FILL_COLOR
        elif numeric_value > weapon_high_threshold:
            fill_color = SUMMARY_HIGH_FILL_COLOR
        if not fill_color:
            continue
        worksheet.cell(row=total_row_index, column=column_offset).fill = PatternFill(
            fill_type="solid",
            fgColor=fill_color.lstrip("#"),
        )


def build_pivot_section(
    result_df: pd.DataFrame,
    section_title: str,
    weapon_column_name: str,
    hero_alert_threshold: float = HERO_USAGE_RATE_LOW_THRESHOLD,
    column_alert_threshold: float = WEAPON_USAGE_RATE_LOW_THRESHOLD,
    block_definitions: tuple[dict[str, Any], ...] = DEFAULT_BLOCK_DEFINITIONS,
    column_sort_key: Callable[[str], tuple[Any, ...]] | None = build_weapon_sort_key,
    side_blocks_by_anchor_title: dict[str, list[dict[str, Any]]] | None = None,
    extra_tables: dict[str, pd.DataFrame] | None = None,
) -> dict[str, Any]:
    # section 是一组完整透视区块的配置对象，Excel 和飞书都复用这份结构。
    pivot_tables = build_hero_weapon_pivot_tables(
        result_df,
        weapon_column_name,
        column_sort_key=column_sort_key,
    )
    section = {
        "section_title": section_title,
        "hero_alert_threshold": hero_alert_threshold,
        "column_alert_threshold": column_alert_threshold,
        "block_definitions": block_definitions,
        "side_blocks_by_anchor_title": side_blocks_by_anchor_title or {},
    }
    section.update(pivot_tables)
    if extra_tables:
        section.update(extra_tables)
    return section


def build_report_sections(
    pivot_source_df: pd.DataFrame,
    weapon_usage_delta_df: pd.DataFrame,
    weapon_evacuation_delta_df: pd.DataFrame,
) -> list[dict[str, Any]]:
    # 五组透视都走同一套 section 结构，便于 Excel 和飞书复用同一份配置。
    weapon_section_block_definitions = tuple(
        {
            **block_definition,
            "color_mode": (
                "weapon_usage_row"
                if block_definition["title"] == "使用率"
                else "weapon_evacuation_row"
                if block_definition["title"] == "撤离率"
                else block_definition.get("color_mode", "global_single")
            ),
        }
        for block_definition in WEAPON_SECTION_BLOCK_DEFINITIONS
    )
    return [
        build_pivot_section(
            pivot_source_df,
            section_title=WEAPON_SECTION_TITLE,
            weapon_column_name="b_main_weapon_type",
            hero_alert_threshold=HERO_USAGE_RATE_LOW_THRESHOLD,
            column_alert_threshold=WEAPON_USAGE_RATE_LOW_THRESHOLD,
            block_definitions=weapon_section_block_definitions,
            side_blocks_by_anchor_title={
                "使用率": [
                    {
                        "title": WEAPON_USAGE_DELTA_BLOCK_TITLE,
                        "data_key": "weapon_usage_delta_df",
                        "percentage": True,
                        "highlight_low_usage": False,
                        "color_mode": "delta_diverging",
                    }
                ],
                "撤离率": [
                    {
                        "title": WEAPON_EVACUATION_DELTA_BLOCK_TITLE,
                        "data_key": "weapon_evacuation_delta_df",
                        "percentage": True,
                        "highlight_low_usage": False,
                        "color_mode": "delta_diverging",
                    }
                ],
            },
            extra_tables={
                "weapon_usage_delta_df": weapon_usage_delta_df,
                "weapon_evacuation_delta_df": weapon_evacuation_delta_df,
            },
        ),
        build_pivot_section(
            pivot_source_df,
            section_title=EQUIPMENT_SET_SECTION_TITLE,
            weapon_column_name="b_equipment_set",
            hero_alert_threshold=HERO_USAGE_RATE_LOW_THRESHOLD,
            column_alert_threshold=ATTACK_PART_USAGE_RATE_ALERT_THRESHOLD,
            block_definitions=SET_SECTION_BLOCK_DEFINITIONS,
            column_sort_key=None,
        ),
        build_pivot_section(
            pivot_source_df,
            section_title=ATTACK_PART_SECTION_TITLE,
            weapon_column_name="b_main_weapon",
            hero_alert_threshold=HERO_USAGE_RATE_LOW_THRESHOLD,
            column_alert_threshold=ATTACK_PART_USAGE_RATE_ALERT_THRESHOLD,
        ),
        build_pivot_section(
            pivot_source_df,
            section_title=WEAPON_ULT_SECTION_TITLE,
            weapon_column_name="b_main_weapon_with_ult",
            hero_alert_threshold=HERO_USAGE_RATE_LOW_THRESHOLD,
            column_alert_threshold=ATTACK_PART_USAGE_RATE_ALERT_THRESHOLD,
            block_definitions=SET_SECTION_BLOCK_DEFINITIONS,
        ),
        build_pivot_section(
            pivot_source_df,
            section_title=WEAPON_BOOST_SECTION_TITLE,
            weapon_column_name="b_main_weapon_boost",
            hero_alert_threshold=HERO_USAGE_RATE_LOW_THRESHOLD,
            column_alert_threshold=ATTACK_PART_USAGE_RATE_ALERT_THRESHOLD,
            block_definitions=SET_SECTION_BLOCK_DEFINITIONS,
            column_sort_key=None,
        ),
    ]


def build_attachment_summary_row(pivot_source_df: pd.DataFrame) -> list[Any]:
    attack_part_scope_df = pivot_source_df.loc[pivot_source_df["b_main_weapon_level"] > 3]
    boost_scope_df = pivot_source_df.loc[pivot_source_df["b_main_weapon_level"] > 4]
    ult_scope_df = pivot_source_df.loc[pivot_source_df["b_main_weapon_level"] > 5]

    attack_part_rate = (
        float(attack_part_scope_df["b_main_weapon_variant"].ne("").mean())
        if not attack_part_scope_df.empty
        else 0.0
    )
    ult_rate = (
        float(ult_scope_df["b_main_weapon_ult"].ne("").mean())
        if not ult_scope_df.empty
        else 0.0
    )
    boost_rate = (
        float(boost_scope_df["b_main_weapon_boost"].ne("").mean())
        if not boost_scope_df.empty
        else 0.0
    )

    return [
        "总量汇总",
        "攻击部比例",
        round(attack_part_rate, 4),
        "奥义比例",
        round(ult_rate, 4),
        "配件比例",
        round(boost_rate, 4),
    ]


def write_attachment_summary_row(
    worksheet: openpyxl.worksheet.worksheet.Worksheet,
    start_row: int,
    summary_row: list[Any],
) -> int:
    for column_index, value in enumerate(summary_row, start=1):
        cell = worksheet.cell(row=start_row, column=column_index, value=value)
        if column_index in (3, 5, 7) and value is not None and not pd.isna(value):
            cell.number_format = "0%"

    return start_row


def write_pivot_summary_section(
    worksheet: openpyxl.worksheet.worksheet.Worksheet,
    start_row: int,
    section: dict[str, Any],
) -> int:
    worksheet.cell(row=start_row, column=1, value=section["section_title"])

    current_row = start_row + 1
    section_last_row = start_row

    for block_definition in section["block_definitions"]:
        _, block_last_row, block_last_column = write_pivot_block(
            worksheet,
            start_row=current_row,
            start_column=1,
            title=block_definition["title"],
            pivot_df=section[block_definition["data_key"]],
            percentage=block_definition["percentage"],
        )
        apply_excel_fill_to_block(
            worksheet,
            section[block_definition["data_key"]],
            start_row=current_row,
            start_column=1,
            block_definition=block_definition,
        )
        if block_definition["highlight_low_usage"]:
            apply_usage_rate_summary_fill_to_excel(
                worksheet,
                section[block_definition["data_key"]],
                start_row=current_row,
                start_column=1,
                hero_low_threshold=HERO_USAGE_RATE_LOW_THRESHOLD,
                hero_high_threshold=HERO_USAGE_RATE_HIGH_THRESHOLD,
                weapon_low_threshold=WEAPON_USAGE_RATE_LOW_THRESHOLD,
                weapon_high_threshold=WEAPON_USAGE_RATE_HIGH_THRESHOLD,
            )
        section_last_row = block_last_row

        side_start_column = block_last_column + SECTION_SIDE_BY_SIDE_GAP_COLUMNS
        for side_block_definition in section.get("side_blocks_by_anchor_title", {}).get(
            block_definition["title"],
            [],
        ):
            _, side_last_row, side_last_column = write_pivot_block(
                worksheet,
                start_row=current_row,
                start_column=side_start_column,
                title=side_block_definition["title"],
                pivot_df=section[side_block_definition["data_key"]],
                percentage=side_block_definition["percentage"],
            )
            apply_excel_fill_to_block(
                worksheet,
                section[side_block_definition["data_key"]],
                start_row=current_row,
                start_column=side_start_column,
                block_definition=side_block_definition,
            )
            section_last_row = max(section_last_row, side_last_row)
            side_start_column = side_last_column + SECTION_SIDE_BY_SIDE_GAP_COLUMNS

        current_row = block_last_row + 3

    return section_last_row


def save_combined_pivot_summary_excel(
    excel_file_path: Path,
    sheet_name: str,
    sections: list[dict[str, Any]],
    attachment_summary_row: list[Any] | None = None,
) -> Path:
    actual_output_path = build_available_output_path(excel_file_path)
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name

    current_row = 1
    for section in sections:
        section_last_row = write_pivot_summary_section(worksheet, current_row, section)
        current_row = section_last_row + 3

    if attachment_summary_row:
        write_attachment_summary_row(worksheet, current_row, attachment_summary_row)

    style_summary_sheet(worksheet)
    workbook.save(actual_output_path)
    return actual_output_path


def upload_pivot_section_to_feishu_sheet(
    access_token: str,
    spreadsheet_token: str,
    sheet_id: str,
    start_row: int,
    section: dict[str, Any],
) -> tuple[int, int, list[str], list[dict[str, Any]], dict[str, str]]:
    # 飞书写入按区块处理，方便把多组透视连续拼接在同一个页签中。
    section_title_range = build_value_range(sheet_id, start_row, [[section["section_title"]]], start_column=1)
    value_ranges = [section_title_range]
    percentage_ranges: list[str] = []
    block_configs: list[dict[str, Any]] = []
    max_columns = 1
    current_row = start_row + 1
    section_last_row = start_row

    for block_definition in section["block_definitions"]:
        block_values = build_pivot_block_values(
            block_definition["title"],
            section[block_definition["data_key"]],
            percentage=block_definition["percentage"],
        )
        value_ranges.append(build_value_range(sheet_id, current_row, block_values, start_column=1))
        main_last_column = len(block_values[1])
        max_columns = max(max_columns, main_last_column)
        if block_definition["percentage"]:
            percentage_ranges.append(
                f"{sheet_id}!B{current_row + 2}:{build_excel_column_name(main_last_column)}{current_row + len(block_values) - 1}"
            )
        block_configs.append(
            {
                "data_frame": section[block_definition["data_key"]],
                "start_row": current_row,
                "start_column": 1,
                "block_definition": block_definition,
                "highlight_low_usage": block_definition["highlight_low_usage"],
                "hero_alert_threshold": section["hero_alert_threshold"],
                "column_alert_threshold": section["column_alert_threshold"],
                "bold_ranges": [
                    f"{sheet_id}!A{current_row}:A{current_row}",
                    (
                        f"{sheet_id}!B{current_row + 1}:"
                        f"{build_excel_column_name(main_last_column)}"
                        f"{current_row + 1}"
                    ),
                    (
                        f"{sheet_id}!A{current_row + 2}:"
                        f"A{current_row + len(block_values) - 1}"
                    ),
                ],
            }
        )

        side_start_column = main_last_column + SECTION_SIDE_BY_SIDE_GAP_COLUMNS
        for side_block_definition in section.get("side_blocks_by_anchor_title", {}).get(
            block_definition["title"],
            [],
        ):
            side_block_values = build_pivot_block_values(
                side_block_definition["title"],
                section[side_block_definition["data_key"]],
                percentage=side_block_definition["percentage"],
            )
            value_ranges.append(
                build_value_range(
                    sheet_id,
                    current_row,
                    side_block_values,
                    start_column=side_start_column,
                )
            )
            side_last_column = side_start_column + len(side_block_values[1]) - 1
            max_columns = max(max_columns, side_last_column)
            if side_block_definition["percentage"]:
                percentage_ranges.append(
                    f"{sheet_id}!{build_excel_column_name(side_start_column + 1)}{current_row + 2}:"
                    f"{build_excel_column_name(side_last_column)}{current_row + len(side_block_values) - 1}"
                )
            block_configs.append(
                {
                    "data_frame": section[side_block_definition["data_key"]],
                    "start_row": current_row,
                    "start_column": side_start_column,
                    "block_definition": side_block_definition,
                    "highlight_low_usage": side_block_definition["highlight_low_usage"],
                    "hero_alert_threshold": section["hero_alert_threshold"],
                    "column_alert_threshold": section["column_alert_threshold"],
                    "bold_ranges": [
                        (
                            f"{sheet_id}!{build_excel_column_name(side_start_column)}{current_row}:"
                            f"{build_excel_column_name(side_start_column)}{current_row}"
                        ),
                        (
                            f"{sheet_id}!{build_excel_column_name(side_start_column + 1)}{current_row + 1}:"
                            f"{build_excel_column_name(side_last_column)}{current_row + 1}"
                        ),
                        (
                            f"{sheet_id}!{build_excel_column_name(side_start_column)}{current_row + 2}:"
                            f"{build_excel_column_name(side_start_column)}{current_row + len(side_block_values) - 1}"
                        ),
                    ],
                }
            )
            side_start_column = side_last_column + SECTION_SIDE_BY_SIDE_GAP_COLUMNS

        section_last_row = current_row + len(block_values) - 1
        current_row = section_last_row + 3

    upload_value_ranges_in_batches(
        access_token,
        spreadsheet_token,
        value_ranges,
        progress_label=section["section_title"],
    )

    return (
        section_last_row,
        max_columns,
        percentage_ranges,
        block_configs,
        {
            "title": section["section_title"],
            "range": f"{sheet_id}!A{start_row}:A{start_row}",
        },
    )


def sync_combined_pivot_tables_to_feishu(
    sections: list[dict[str, Any]],
    report_name: str,
    attachment_summary_row: list[Any] | None = None,
) -> str:
    access_token = get_feishu_access_token()
    spreadsheet_token = get_feishu_sheet_token(access_token, FEISHU_REPORT_WIKI_URL)
    sheet_title = build_feishu_sheet_title(report_name)
    try:
        sheet_id = create_feishu_sheet(access_token, spreadsheet_token, sheet_title)
    except FeishuApiError as exc:
        if "sheetTitle already exist" not in str(exc):
            raise
        sheet_title = f"{sheet_title}_{datetime.now().strftime('%H%M%S')}"
        sheet_id = create_feishu_sheet(access_token, spreadsheet_token, sheet_title)

    current_row = 1
    max_columns = 1
    last_row = 1
    percentage_ranges: list[str] = []
    block_configs: list[dict[str, Any]] = []
    section_title_infos: list[dict[str, str]] = []

    for section in sections:
        section_last_row, section_max_columns, section_percentage_ranges, section_block_configs, section_title_info = (
            upload_pivot_section_to_feishu_sheet(
                access_token,
                spreadsheet_token,
                sheet_id,
                current_row,
                section,
            )
        )
        last_row = section_last_row
        max_columns = max(max_columns, section_max_columns)
        percentage_ranges.extend(section_percentage_ranges)
        block_configs.extend(section_block_configs)
        section_title_infos.append(section_title_info)
        current_row = section_last_row + 3

    summary_percentage_ranges: list[str] = []
    summary_bold_range = ""
    if attachment_summary_row:
        summary_value_range = build_value_range(sheet_id, current_row, [attachment_summary_row])
        upload_value_ranges_in_batches(
            access_token,
            spreadsheet_token,
            [summary_value_range],
            progress_label="挂件汇总",
        )
        last_row = current_row
        max_columns = max(max_columns, len(attachment_summary_row))
        summary_bold_range = f"{sheet_id}!A{current_row}:G{current_row}"
        summary_percentage_ranges.extend(
            [
                f"{sheet_id}!C{current_row}:C{current_row}",
                f"{sheet_id}!E{current_row}:E{current_row}",
                f"{sheet_id}!G{current_row}:G{current_row}",
            ]
        )

    full_range = f"{sheet_id}!A1:{build_excel_column_name(max_columns)}{last_row}"
    style_requests: list[dict[str, Any]] = [
        build_alignment_style_request([full_range]),
    ]

    for section_title_info in section_title_infos:
        style_requests.append(build_bold_style_request([section_title_info["range"]]))
        if section_title_info["title"] == WEAPON_SECTION_TITLE:
            style_requests.append(build_left_alignment_style_request([section_title_info["range"]]))

    for percentage_range in percentage_ranges:
        style_requests.append(build_percentage_style_request([percentage_range]))
    for percentage_range in summary_percentage_ranges:
        style_requests.append(build_percentage_style_request([percentage_range]))
    if summary_bold_range:
        style_requests.append(build_bold_style_request([summary_bold_range]))

    for block_config in block_configs:
        try:
            for bold_range in block_config["bold_ranges"]:
                style_requests.append(build_bold_style_request([bold_range]))
            style_requests.extend(
                apply_feishu_fill_to_block(
                    sheet_id,
                    block_config["data_frame"],
                    block_config["start_row"],
                    block_config["start_column"],
                    block_config["block_definition"],
                )
            )
        except Exception as exc:
            logger.warning("飞书条件色阶设置失败：%s", exc)

    for block_config in block_configs:
        if not block_config.get("highlight_low_usage"):
            continue

        try:
            usage_summary_style_requests = build_usage_rate_summary_style_requests(
                sheet_id,
                block_config["data_frame"],
                block_config["start_row"],
                block_config["start_column"],
                HERO_USAGE_RATE_LOW_THRESHOLD,
                HERO_USAGE_RATE_HIGH_THRESHOLD,
                WEAPON_USAGE_RATE_LOW_THRESHOLD,
                WEAPON_USAGE_RATE_HIGH_THRESHOLD,
            )
            if usage_summary_style_requests:
                style_requests.extend(usage_summary_style_requests)
        except Exception as exc:
            logger.warning("飞书低使用率高亮设置失败：%s", exc)

    try:
        batch_update_feishu_styles(
            access_token,
            spreadsheet_token,
            style_requests,
        )
    except Exception as exc:
        logger.warning("飞书样式批量设置失败：%s", exc)

    try:
        update_feishu_dimension_range(
            access_token,
            spreadsheet_token,
            dimension={
                "sheetId": sheet_id,
                "majorDimension": "COLUMNS",
                "startIndex": 1,
                "endIndex": max_columns,
            },
            dimension_properties={"visible": True, "fixedSize": 80},
        )
    except Exception as exc:
        logger.warning("飞书列宽设置失败：%s", exc)

    return sheet_title


def fetch_battle_data(query_date_value: date) -> dict[str, Any]:
    """拉取原始对战数据和映射配置。"""
    set_runtime_query_context(query_date_value)
    logger.info("开始拉取数据，查询日期：%s", QUERY_DATE)
    logger.info("映射文件：%s", ITEM_INFO_EXCEL_PATH)
    logger.info("周基准缓存：%s", WEEKLY_BASELINE_CACHE_PATH)
    logger.info("版本更新时间配置：%s", VERSION_UPDATE_TIME_PATH)

    column_names, rows = execute_query(
        build_sql_query([QUERY_DATE], time_lower_bound=CURRENT_PERIOD_LOWER_BOUND)
    )
    id_name_mapping = load_item_name_mapping(ITEM_INFO_EXCEL_PATH)
    equipment_name_to_set_mapping, set_name_candidates = load_equipment_set_reference(
        ITEM_INFO_EXCEL_PATH
    )
    logger.info("数据拉取完成，共获取 %s 行原始数据。", len(rows))

    return {
        "column_names": column_names,
        "rows": rows,
        "id_name_mapping": id_name_mapping,
        "equipment_name_to_set_mapping": equipment_name_to_set_mapping,
        "set_name_candidates": set_name_candidates,
    }


def process_battle_data(raw_data: dict[str, Any], query_date_value: date) -> dict[str, Any]:
    """处理原始数据，生成透视统计和基准变化结果。"""
    logger.info("开始处理数据并构建透视报表。")
    result_df = pd.DataFrame(raw_data["rows"], columns=raw_data["column_names"])
    result_df = normalize_weapon_source_columns(result_df)
    translated_result_df = translate_result_dataframe(result_df, raw_data["id_name_mapping"])
    pivot_source_df = prepare_pivot_source_dataframe(
        translated_result_df,
        raw_data["equipment_name_to_set_mapping"],
        raw_data["set_name_candidates"],
    )
    weapon_usage_delta_df, weapon_evacuation_delta_df, baseline_window = (
        build_weapon_baseline_delta_tables(
            pivot_source_df,
            raw_data["equipment_name_to_set_mapping"],
            raw_data["set_name_candidates"],
            query_date_value,
        )
    )
    logger.info(
        "数据处理完成，变化基准周期：%s ~ %s",
        baseline_window["start_date"].strftime("%Y-%m-%d"),
        baseline_window["end_date"].strftime("%Y-%m-%d"),
    )

    attachment_summary_row = build_attachment_summary_row(pivot_source_df)
    sections = build_report_sections(
        pivot_source_df,
        weapon_usage_delta_df,
        weapon_evacuation_delta_df,
    )
    return {
        "rows_count": len(raw_data["rows"]),
        "sections": sections,
        "attachment_summary_row": attachment_summary_row,
    }


def export_excel_and_sync_feishu(processed_data: dict[str, Any]) -> dict[str, Any]:
    """导出 Excel 报表，并同步结果到飞书。"""
    output_path = save_combined_pivot_summary_excel(
        build_output_excel_path(""),
        REPORT_WORKSHEET_TITLE,
        processed_data["sections"],
        attachment_summary_row=processed_data["attachment_summary_row"],
    )
    logger.info("Excel 导出完成：%s", output_path)

    feishu_sheet_title = ""
    try:
        feishu_sheet_title = sync_combined_pivot_tables_to_feishu(
            processed_data["sections"],
            "",
            attachment_summary_row=processed_data["attachment_summary_row"],
        )
        logger.info("飞书同步完成，页签：%s", feishu_sheet_title)
    except FeishuApiError as exc:
        logger.warning("飞书同步失败：%s", exc)

    return {
        "output_path": output_path,
        "feishu_sheet_title": feishu_sheet_title,
    }


def run_report_job() -> None:
    query_date_value = resolve_query_date()
    logger.info("开始执行单局对战信息任务。")
    try:
        raw_data = fetch_battle_data(query_date_value)
        processed_data = process_battle_data(raw_data, query_date_value)
        export_result = export_excel_and_sync_feishu(processed_data)
        logger.info("任务执行完成，共导出 %s 行数据。", processed_data["rows_count"])
        logger.info("结果已保存：%s", export_result["output_path"])
        if export_result["feishu_sheet_title"]:
            logger.info("飞书页签已创建并同步：%s", export_result["feishu_sheet_title"])
    except Exception:
        logger.exception("任务执行失败。")
        raise


def main() -> None:
    setup_logging()
    runtime_config = load_runtime_config()
    apply_runtime_config(runtime_config)

    logger.info("配置加载完成。")
    logger.info("飞书同步链接：%s", FEISHU_REPORT_WIKI_URL)
    logger.info("日志文件：%s", LOG_FILE_PATH)
    run_report_job()


if __name__ == "__main__":
    main()



